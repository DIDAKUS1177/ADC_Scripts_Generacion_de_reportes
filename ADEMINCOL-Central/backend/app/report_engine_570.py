"""
Generador REAL de reportes API 570 (Inspección Visual de Tubería) usando la
plantilla verificada (templates_xlsx/570.xlsx) y datos reales del Sheet
`DB_INSP_API_570`. Mismo patrón que report_engine_mt.py, generalizado a 15
secciones independientes (cada una con su propia tabla de datos + bloque de
fotos), tal como está diseñado el script GAS original
(APP011_Tub_570_VT/APP011_Tub_570_VT.js — SECTIONS_CONFIG).

Diferencia clave frente al GAS original: aquí cada sección se procesa
COMPLETA (datos + fotos) antes de pasar a la siguiente, acumulando un único
offset de filas insertadas. El script GAS separaba "loop de datos" y "loop de
fotos" (por límites de tiempo de Apps Script, que no aplican aquí) y por eso
el cálculo de `basePhotoRow` de una sección no consideraba las filas de fotos
insertadas por secciones ANTERIORES en el loop de fotos — un desalineamiento
potencial que no se replica en este motor.
"""
import io
import logging
import re
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .image_utils import desactivar_fit_to_page, descargar_imagen, insertar_imagen_centrada
from .report_utils import valor_tipado

logger = logging.getLogger("report_engine_570")

TEMPLATE_PATH = Path(__file__).resolve().parent / "templates_xlsx" / "570.xlsx"
HOJA_FORMATO = "FORMATO"

CELDAS_GENERALES = {
    "cliente": "I8", "consecutivo": "V8", "fecha": "AI8", "ubicacion": "AV8",
    "ot": "I10", "servicio": "V10", "codigo_fabricacion": "AI10", "ano_fabricacion": "AV10",
    "sistema": "I12", "subsistema": "V12", "presion_operacion": "AI12", "temperatura_operacion": "AV12",
}
CELDA_FIRMA = "J165"
CELDAS_FIRMA_TEXTO = {"nombre": "J166", "cargo": "J167", "certificacion": "J168", "fecha": "J169"}

# Bloques "REVISADO POR" (columna Y) y "APROBADO POR" (columna AN), fila
# 164-169 — mismas filas que "REALIZADO POR" (columna J), verificado contra
# la plantilla real el 2026-07-14 (Y165:AM165 y AN165:BB165 son celdas
# combinadas vacías en la plantilla, listas para llenarse). Pedido explícito
# del usuario: dar libertad de elegir quién revisa/aprueba cada reporte
# (usuario registrado en la plataforma o datos manuales) — resuelto en
# main.py (_generar_bytes_570) contra 'revisor_*'/'aprobador_*' en
# fila_general. La fecha de cada bloque es SIEMPRE automática (hoy), nunca
# viene de overrides.
CELDA_FIRMA_REVISOR = "Y165"
CELDAS_TEXTO_REVISOR = {"nombre": "Y166", "cargo": "Y167", "certificado": "Y168", "fecha": "Y169"}
CELDA_FIRMA_APROBADOR = "AN165"
CELDAS_TEXTO_APROBADOR = {"nombre": "AN166", "cargo": "AN167", "certificado": "AN168", "fecha": "AN169"}

# Traducido literal de SECTIONS_CONFIG en APP011_Tub_570_VT.js (2026-07-03)
SECTIONS_CONFIG = {
    "coating": {
        "sheet": "#2_recubrimiento", "photo_sheet": "#2_recubrimiento_photos", "data_start_row": 17,
        "mapping": {"segmento_linea": "C", "cml_tml": "F", "nps": "H", "tipo_componente": "J", "tipo_dano": "O", "tipo_calidad": "S", "area_reparacion": "W", "observaciones_recubrimiento": "AB"},
        "photo_row": 21, "desc_row": 22, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "interfaces": {
        "sheet": "#3_interfases_2", "photo_sheet": "#3_interfasesueloairepenetraciondeparedes_photos", "data_start_row": 27,
        "mapping": {"segmento_linea_interfase": "C", "cml_tml_interfase": "F", "nps_interfase": "H", "tipo_componente_interfase": "J", "estado_recubrimiento": "O", "observaciones_interfase": "W"},
        "photo_row": 31, "desc_row": 32, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "supports": {
        "sheet": "#4_soportes", "photo_sheet": "#4_soportes_photos", "data_start_row": 37,
        "mapping": {"segmento_linea_soporte": "C", "cml_tml_soporte": "F", "nps_soporte": "H", "tipo_componente_soporte": "J", "id_soporte": "M", "tipo_soporte": "O", "anclaje_soporte": "R", "accesorio_soporte": "U", "aislamiento_soporte": "X", "contacto_soporte": "AA", "estado_recubrimiento_soporte": "AD", "estado_concreto_soporte": "AG", "ausencia_partes_soporte": "AJ", "desajuste_partes_soporte": "AM", "corrosion_soporte": "AP", "deformacion_soporte": "AS", "observaciones_soporte": "AV"},
        "photo_row": 41, "desc_row": 42, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "vibration": {
        "sheet": "#5_vibracion", "photo_sheet": "#5_vibracion_photos", "data_start_row": 47,
        "mapping": {"segmento_linea_vibracion": "C", "cml_tml_vibracion": "F", "nps_vibracion": "H", "tipo_componente_vibracion": "J", "condicion_vibracion": "O", "fuente_vibracion": "S", "punto_friccion": "X", "observaciones_vibracion": "AC"},
        "photo_row": 51, "desc_row": 52, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "deadLegs": {
        "sheet": "#6_piernasmuertas", "photo_sheet": "#6_piernasmuertas_photos", "data_start_row": 57,
        "mapping": {"segmento_linea_pm": "C", "cml_tml_pm": "F", "nps_pm": "H", "tipo_componente_pm": "J", "id_pierna_muerta": "O", "longitud_pierna_muerta": "R", "posicion_pierna_muerta": "V", "observaciones_pm": "Z"},
        "photo_row": 61, "desc_row": 62, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "reliefDevices": {
        "sheet": "#7_dispositivos", "photo_sheet": "#7_dispositivosdealiviodepresion_photos", "data_start_row": 67,
        "mapping": {"segmento_linea_disp_alivio": "C", "cml_tml_disp_alivio": "F", "tag_disp_alivio": "H", "marca_disp_alivio": "K", "modelo_disp_alivio": "N", "serial_disp_alivio": "Q", "tamano_entrada_disp_alivio": "T", "tamano_salida_disp_alivio": "W", "fecha_calibracion_disp_alivio": "Z", "presion_calibracion_disp_alivio": "AC", "fugas_bridas_pernos": "AF", "danos_recubrimiento_disp_alivio": "AI", "seguridad_pernos_bridas": "AL", "precintos_valvulas_corte": "AP", "observaciones_disp_alivio": "AU"},
        "photo_row": 71, "desc_row": 72, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "mechCorrosion": {
        "sheet": "#8_estadomecanicoycorrosion", "photo_sheet": "#8_estadomecanicoycorrosion_photos", "data_start_row": 77,
        "mapping": {"segmento_linea_emc": "C", "cml_tml_emc": "F", "nps_emc": "H", "tipo_componente_emc": "J", "condicion_emc": "O", "observaciones_emc": "V"},
        "photo_row": 81, "desc_row": 82, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "flangeJoints": {
        "sheet": "#9_unionesbridadas", "photo_sheet": "#9_unionesbridadas_photos", "data_start_row": 87,
        "mapping": {"segmento_linea_brida": "C", "cml_tml_brida": "F", "nps_brida": "H", "tipo_componente_brida": "J", "tipo_brida": "O", "rating_class_brida": "R", "tipo_cara_brida": "U", "llenado_tuerca_brida": "X", "estado_recubrimiento_brida": "AA", "presenta_fugas_brida": "AE", "presenta_junta_disimil": "AH", "observaciones_brida": "AK"},
        "photo_row": 91, "desc_row": 92, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "valves": {
        "sheet": "#10_valvulas", "photo_sheet": "#10_valvulasdecorteyunidireccionales_photos", "data_start_row": 97,
        "mapping": {"segmento_linea_valvula": "C", "cml_tml_valvula": "F", "nps_valvula": "H", "tipo_componente_valvula": "J", "tipo_de_valvula": "O", "material_cuerpo_valvula": "R", "rating_class_valvula": "U", "tipo_conexion_extremos": "X", "condicion_sello_valvula": "AA", "estado_recubrimiento_valvula": "AD", "observaciones_valvula": "AH"},
        "photo_row": 101, "desc_row": 102, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "instruments": {
        "sheet": "#11_instrumentos", "photo_sheet": "#11_instrumentos_photos", "data_start_row": 107,
        "mapping": {"segmento_linea_instrumento": "C", "cml_tml_instrumento": "F", "nps_instrumento": "H", "tipo_de_instrumento": "J", "tag_instrumento": "O", "observaciones_instrumento": "S"},
        "photo_row": 111, "desc_row": 112, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "corrosionCoupons": {
        "sheet": "#12_cuponesdecorrosion", "photo_sheet": "#12_cuponesdecorrosion_photos", "data_start_row": 117,
        "mapping": {"segmento_linea_cupon": "C", "cml_tml_cupon": "F", "nps_cupon": "H", "tipo_de_cupon": "J", "tag_cupon": "O", "observaciones_cupon": "S"},
        "photo_row": 121, "desc_row": 122, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "injectionPoints": {
        "sheet": "#13_puntosdeinyeccion", "photo_sheet": "#13_puntosdeinyeccion_photos", "data_start_row": 127,
        "mapping": {"segmento_linea_inyeccion": "C", "cml_tml_inyeccion": "F", "nps_inyeccion": "H", "no_punto_inyeccion": "J", "estado_mecanico_inyeccion": "N"},
        "photo_row": 131, "desc_row": 132, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "thermalInsulation": {
        "sheet": "#14_aislamientotermico", "photo_sheet": "#14_aislamientotermico_photos", "data_start_row": 137,
        "mapping": {"segmento_linea_aislamiento": "C", "cml_tml_aislamiento": "F", "nps_aislamiento": "H", "danos_perforaciones_aislamiento": "J", "falta_recubrimiento_aislamiento": "O", "deterioro_sellado_aislamiento": "U", "abultamiento_aislamiento": "Y", "cintas_rotas_faltantes": "AC", "observaciones_aislamiento": "AG"},
        "photo_row": 141, "desc_row": 142, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "corrosionExterna": {
        "sheet": "#15_corrocion_externa", "photo_sheet": "#15_corrocion_externa_photos", "data_start_row": 143,
        "mapping": {"segmento_linea": "C", "cml_tml": "F", "nps": "H", "fugas_superficie_externa": "J", "aplastamientos_ovalidad": "N", "estrias_hendiduras_cortes": "R", "profundidad_estrias": "V", "buena_fusion_uniones": "Z", "contaminacion_uniones": "AD", "grietas_agrietamiento_crazing": "AH", "fugas_uniones_accesorios": "AL", "observaciones": "AP"},
        "photo_row": 151, "desc_row": 152, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "polimeros": {
        "sheet": "#16_polimeros", "photo_sheet": "#16_polimeros_photos", "data_start_row": 157,
        "mapping": {"tipo_componente": "C", "presencia_ampollas": "G", "diametro_blister": "J", "longitud_microfisura": "M", "perdida_resina": "O", "aranazos_hendiduras": "Q", "ancho_aranazo": "T", "longitud_aranazo": "V", "decoloracion_superficie": "X", "color_decoloracion": "Z", "grietas_fracturas": "AB", "inclusiones": "AD", "quemaduras": "AF", "bordes_expuestos": "AH", "delaminaciones": "AK", "arrugas": "AN", "observaciones_externas": "AQ"},
        "photo_row": 161, "desc_row": 162, "photo_cols": ["C", "P", "AC", "AP"],
    },
}

SECTION_KEYS_ORDEN = list(SECTIONS_CONFIG.keys())


def _col_fila(celda: str) -> tuple[str, int]:
    """Separa una referencia de celda en (columna, fila) — a diferencia de
    `celda[0]`/`celda[1:]` (usado en CELDA_FIRMA/CELDAS_FIRMA_TEXTO, columna
    J, siempre de una sola letra), soporta columnas de dos letras como 'AN'
    (bug encontrado 2026-07-14 al agregar el bloque "Aprobado por": 'AN165'
    partido con `celda[0]`/`celda[1:]` daba columna 'A' y "N165", que no es
    un entero)."""
    m = re.match(r"([A-Z]+)(\d+)", celda)
    return m.group(1), int(m.group(2))


def _escribir_texto_negro(ws, celda: str, valor):
    """Escribe un valor y fuerza fuente negra — las celdas de los bloques
    revisor/aprobador (Y166-169/AN166-169) heredan en la plantilla el color
    blanco del encabezado rojo de arriba (theme=0), lo que las deja
    invisibles sobre el fondo blanco/gris de esas filas (bug reportado por
    el usuario 2026-07-16, texto "en blanco" aunque el dato sí estaba)."""
    ws[celda] = valor
    f = ws[celda].font
    ws[celda].font = Font(
        name=f.name, size=f.size, bold=f.bold, italic=f.italic, color="FF000000"
    )


def _insertar_filas_y_ajustar_alturas(ws, pos: int, n: int):
    """Igual que en report_engine_mt.py: openpyxl `insert_rows()` no desplaza
    alturas de fila ni merges — se hace manualmente aquí."""
    if n <= 0:
        return
    max_row_antes = ws.max_row
    alturas_originales = {
        r: ws.row_dimensions[r].height
        for r in range(pos, max_row_antes + 1)
        if r in ws.row_dimensions and ws.row_dimensions[r].height is not None
    }
    merges_a_mover = [r for r in list(ws.merged_cells.ranges) if r.min_row >= pos]
    for rng in merges_a_mover:
        ws.unmerge_cells(str(rng))

    ws.insert_rows(pos, n)

    for rng in merges_a_mover:
        rng.shift(0, n)
        ws.merge_cells(str(rng))

    for r in range(pos, max_row_antes + n + 1):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    for r_original, altura in alturas_originales.items():
        ws.row_dimensions[r_original + n].height = altura


def _replicar_merges_de_fila(ws, fila_origen: int, fila_destino: int):
    origen_merges = [
        rng for rng in list(ws.merged_cells.ranges)
        if rng.min_row == fila_origen and rng.max_row == fila_origen
    ]
    for rng in origen_merges:
        ref = (
            f"{get_column_letter(rng.min_col)}{fila_destino}:"
            f"{get_column_letter(rng.max_col)}{fila_destino}"
        )
        try:
            ws.merge_cells(ref)
        except ValueError:
            pass


def _copiar_estilo_fila(ws, fila_origen: int, fila_destino: int, max_col: int = 56):
    for c in range(1, max_col + 1):
        origen = ws.cell(row=fila_origen, column=c)
        destino = ws.cell(row=fila_destino, column=c)
        destino.font = copy(origen.font)
        destino.border = copy(origen.border)
        destino.fill = copy(origen.fill)
        destino.alignment = copy(origen.alignment)
        destino.number_format = origen.number_format


def generar_reporte_570(
    fila_general: dict,
    secciones_data: dict[str, list[dict]],
    secciones_fotos: dict[str, list[dict]],
    progreso=None,
) -> bytes:
    """Genera el .xlsx real de API 570 y devuelve los bytes.

    `secciones_data[key]` = lista de dicts (filas de la hoja de esa sección
    filtradas por id_api570). `secciones_fotos[key]` = lista de dicts
    {url, descripcion} para esa sección.

    A diferencia del script GAS original, cada sección se procesa completa
    (datos + fotos) antes de pasar a la siguiente, con un único offset
    acumulado — evita el desalineamiento potencial de fotos en secciones
    posteriores que tiene el GAS (ver docstring del módulo).
    """
    def _reportar(pct: int, etapa: str):
        if progreso:
            progreso(pct, etapa)

    _reportar(3, "Preparando plantilla")
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[HOJA_FORMATO]
    desactivar_fit_to_page(ws)

    _reportar(5, "Escribiendo datos generales")
    for campo, celda in CELDAS_GENERALES.items():
        valor = fila_general.get(campo)
        if valor:
            ws[celda] = valor_tipado(valor)

    total_fotos = sum(len(f) for f in secciones_fotos.values()) or 1
    fotos_procesadas = 0
    filas_acumuladas = 0

    for idx_sec, key in enumerate(SECTION_KEYS_ORDEN):
        config = SECTIONS_CONFIG[key]
        registros = secciones_data.get(key, [])
        fotos = secciones_fotos.get(key, [])

        pct_base = 10 + round((idx_sec / len(SECTION_KEYS_ORDEN)) * 80)
        _reportar(pct_base, f"Sección {idx_sec + 1}/{len(SECTION_KEYS_ORDEN)}: {config['sheet']}")

        fila_inicio = config["data_start_row"] + filas_acumuladas
        filas_extra_datos = max(0, len(registros) - 2)  # plantilla trae 2 filas de capacidad

        if filas_extra_datos > 0:
            fila_patron = fila_inicio + 1
            _insertar_filas_y_ajustar_alturas(ws, fila_patron + 1, filas_extra_datos)
            altura_patron = ws.row_dimensions[fila_patron].height
            for i in range(filas_extra_datos):
                fila_nueva = fila_patron + 1 + i
                _copiar_estilo_fila(ws, fila_patron, fila_nueva)
                _replicar_merges_de_fila(ws, fila_patron, fila_nueva)
                if altura_patron:
                    ws.row_dimensions[fila_nueva].height = altura_patron

        for i, reg in enumerate(registros):
            fila_actual = fila_inicio + i
            for campo, col in config["mapping"].items():
                valor = reg.get(campo)
                if valor:
                    ws[f"{col}{fila_actual}"] = valor_tipado(valor)

        filas_acumuladas += filas_extra_datos

        # ---- Fotos de esta sección ----
        base_photo_row = config["photo_row"] + filas_acumuladas
        base_desc_row = config["desc_row"] + filas_acumuladas

        chunks_necesarios = -(-max(len(fotos), 1) // 4)  # ceil(n/4), mínimo 1
        if chunks_necesarios > 1:
            filas_extra_fotos = (chunks_necesarios - 1) * 2
            _insertar_filas_y_ajustar_alturas(ws, base_photo_row + 1, filas_extra_fotos)
            for c in range(1, chunks_necesarios):
                f_foto = base_photo_row + c * 2
                f_desc = f_foto + 1
                _copiar_estilo_fila(ws, base_photo_row, f_foto)
                _copiar_estilo_fila(ws, base_desc_row, f_desc)
                _replicar_merges_de_fila(ws, base_photo_row, f_foto)
                _replicar_merges_de_fila(ws, base_desc_row, f_desc)
                ws.row_dimensions[f_foto].height = ws.row_dimensions[base_photo_row].height
                ws.row_dimensions[f_desc].height = ws.row_dimensions[base_desc_row].height
            filas_acumuladas += filas_extra_fotos

        for i, foto in enumerate(fotos):
            chunk_idx = i // 4
            pos_in_chunk = i % 4
            f_foto = base_photo_row + chunk_idx * 2
            f_desc = f_foto + 1
            col = config["photo_cols"][pos_in_chunk]

            desc = foto.get("descripcion") or ""
            if desc:
                ws[f"{col}{f_desc}"] = desc

            img_bytes = descargar_imagen(foto.get("url") or "")
            if img_bytes:
                insertar_imagen_centrada(ws, img_bytes, f"{col}{f_foto}")

            fotos_procesadas += 1
            pct_fotos = 10 + round((fotos_procesadas / total_fotos) * 80)
            _reportar(min(pct_fotos, 90), f"Foto {fotos_procesadas} de {total_fotos}")

    # La firma va DESPUÉS de las 15 secciones — su fila real se corre hacia
    # abajo por cada fila que se haya insertado dinámicamente arriba (mismo
    # bug encontrado y corregido en report_engine_510.py el 2026-07-03:
    # escribir en una fila fija producía "MergedCell... read-only" en cuanto
    # alguna sección insertaba filas).
    _reportar(93, "Insertando firma")
    fila_firma = int(CELDA_FIRMA[1:]) + filas_acumuladas
    firma_bytes = descargar_imagen(fila_general.get("link_firma", ""))
    if firma_bytes:
        insertar_imagen_centrada(ws, firma_bytes, f"J{fila_firma}", recortar_contenido=True)
    for campo, celda in CELDAS_FIRMA_TEXTO.items():
        valor = fila_general.get(campo)
        if valor:
            col = celda[0]
            fila = int(celda[1:]) + filas_acumuladas
            ws[f"{col}{fila}"] = valor_tipado(valor)

    _reportar(94, "Insertando revisor y aprobador")
    for prefijo, celda_firma, celdas_texto in (
        ("revisor", CELDA_FIRMA_REVISOR, CELDAS_TEXTO_REVISOR),
        ("aprobador", CELDA_FIRMA_APROBADOR, CELDAS_TEXTO_APROBADOR),
    ):
        nombre = fila_general.get(f"{prefijo}_nombre")
        if not nombre:
            continue
        col_nom, fila_nom = _col_fila(celdas_texto["nombre"])
        _escribir_texto_negro(ws, f"{col_nom}{fila_nom + filas_acumuladas}", valor_tipado(nombre))
        for campo in ("cargo", "certificado"):
            valor = fila_general.get(f"{prefijo}_{campo}")
            if valor:
                col_c, fila_c = _col_fila(celdas_texto[campo])
                _escribir_texto_negro(ws, f"{col_c}{fila_c + filas_acumuladas}", valor_tipado(valor))
        col_f, fila_f = _col_fila(celdas_texto["fecha"])
        _escribir_texto_negro(ws, f"{col_f}{fila_f + filas_acumuladas}", datetime.now().strftime("%Y-%m-%d"))
        firma_bytes_bloque = descargar_imagen(fila_general.get(f"{prefijo}_firma_link", ""))
        if firma_bytes_bloque:
            col_fi, fila_fi = _col_fila(celda_firma)
            insertar_imagen_centrada(
                ws, firma_bytes_bloque, f"{col_fi}{fila_fi + filas_acumuladas}", recortar_contenido=True
            )

    _reportar(97, "Guardando archivo")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
