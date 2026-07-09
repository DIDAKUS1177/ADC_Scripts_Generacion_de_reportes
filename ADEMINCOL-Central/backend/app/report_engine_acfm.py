"""
Generador REAL de reportes APP015 Insp ACFM usando la plantilla verificada
(templates_xlsx/ACFM.xlsx) y datos reales del Sheet `ID_BD_GENERAL`
(ReporteACFM.gs, aportado por el usuario 2026-07-09 — el script encontrado
originalmente en la carpeta del proyecto era un stub sin lógica).

Traducido lo más literal posible del GAS original porque el patrón de
inserción de filas es SUTIL y distinto al de 570/Piernas Muertas:

1. **Capacidad de la plantilla = 1 fila, no 2.** `processSection_Excel`
   inserta `cantidadRegistros - 1` filas cuando hay más de un registro (no
   `max(0, n - 2)` como en 570/PM). La plantilla real SÍ tiene una fila 34
   con el mismo patrón de merges que la 33 (aparenta capacidad de 2), pero
   el script nunca la trata como tal — siempre ancla en `dataStartRow` y
   escribe secuencialmente desde ahí. Se replica tal cual: no es un bug a
   corregir, es la lógica original que el usuario pidió respetar.
2. **Dos "secciones", pero la segunda no tiene datos propios.**
   `fotosGenerales` apunta a la hoja general (`1.0_general`) con un
   `mapping` vacío — nunca se leen registros para ella, solo sirve de
   ancla para un SEGUNDO bloque de fotos independiente (`1.1_general_
   PHOTOS`) que no está atado a ningún registro de `1.1_reporte_datos`.
3. **Las fotos de "datosACFM" se filtran por `id_general`, no por el id de
   cada registro individual** (`id_datos`) — aunque la hoja de fotos tiene
   ambas columnas, el GAS solo usa `id_general`. Todas las fotos del PVID
   aparecen en el mismo bloque, sin importar a qué fila de datos
   pertenecen.
4. **Imágenes de la fila general** (esquema/registro/firma) se insertan
   aparte del loop de secciones, en celdas fijas que si acumulan offset por
   filas insertadas arriba.
"""
import io
import logging
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .image_utils import desactivar_fit_to_page, descargar_imagen, insertar_imagen_centrada
from .report_utils import valor_tipado

logger = logging.getLogger("report_engine_acfm")

TEMPLATE_PATH = Path(__file__).resolve().parent / "templates_xlsx" / "ACFM.xlsx"
HOJA_FORMATO = "FORMATO"

# Traducido literal de MAPEO_DE_CELDAS en ReporteACFM.gs
CELDAS_GENERALES = {
    "cliente": "D7", "contrato": "N7", "proyecto": "D9", "ot_n": "N9",
    "no_reporte": "D11", "troncal": "N11", "estacion": "D13", "sistema": "N13",
    "tag": "D15", "capacidad": "N15", "dr_pk": "D17", "fecha": "N17",
    "equipo_acfm": "D21", "tipo_sonda": "N21", "serie": "D23", "n_serie": "N23",
    "fecha_calibracion": "D25", "frecuencia": "N25", "observaciones": "A37",
    "nombre": "C46", "cargo": "C47", "certificado": "C48", "fecha_firma": "C49",
}

# Imágenes de la fila general (fuera de MAPEO_DE_CELDAS en el GAS original,
# insertadas aparte con insertarImagenEnCelda_Excel).
CELDA_ESQUEMA = "A28"
CELDA_REGISTRO = "I28"
CELDA_FIRMA = "C45"

# Traducido literal de SECTIONS_CONFIG en ReporteACFM.gs
SECTIONS_CONFIG = {
    "datosACFM": {
        "sheet": "1.1_reporte_datos", "photo_sheet": "1.1_reporte_datos_PHOTOS", "data_start_row": 33,
        "mapping": {
            "equipo": "A", "segmento": "B", "n_cml": "C", "diametro_pulg": "D",
            "espesor_pulg": "E", "longitud_junta_m": "F", "indicaciones_lado_a": "H",
            "longitud_estimada_mm": "I", "longitud_real_mm": "K", "profundidad_mm": "L",
            "reporte_anexo_grafico_no": "M", "observaciones": "N",
        },
        "photo_row": 39, "desc_row": 40, "photo_cols": ["A", "E", "I", "M"],
    },
    "fotosGenerales": {
        # Sin "sheet"/"mapping" reales: en el GAS original esta sección solo
        # existe para anclar un segundo bloque de fotos independiente de
        # cualquier registro de datosACFM (mapping vacío -> nunca se lee ni
        # inserta fila de datos).
        "photo_sheet": "1.1_general_PHOTOS", "data_start_row": 41,
        "mapping": {},
        "photo_row": 42, "desc_row": 43, "photo_cols": ["A", "E", "I", "M"],
    },
}

SECTION_KEYS_ORDEN = list(SECTIONS_CONFIG.keys())


def _insertar_filas_y_ajustar_alturas(ws, pos: int, n: int):
    """Igual que en report_engine_570.py: openpyxl `insert_rows()` no
    desplaza alturas de fila ni merges — se hace manualmente aquí."""
    if n <= 0:
        return
    max_row_antes = ws.max_row
    alturas_originales = {
        r: ws.row_dimensions[r].height
        for r in range(pos, max_row_antes + 1)
        if r in ws.row_dimensions and ws.row_dimensions[r].height is not None
    }
    merges_a_mover = [r for r in list(ws.merged_cells.ranges) if r.min_row >= pos]
    for rng in merges_a_mover:
        ws.unmerge_cells(str(rng))

    ws.insert_rows(pos, n)

    for rng in merges_a_mover:
        rng.shift(0, n)
        ws.merge_cells(str(rng))

    for r in range(pos, max_row_antes + n + 1):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    for r_original, altura in alturas_originales.items():
        ws.row_dimensions[r_original + n].height = altura


def _replicar_merges_de_fila(ws, fila_origen: int, fila_destino: int):
    origen_merges = [
        rng for rng in list(ws.merged_cells.ranges)
        if rng.min_row == fila_origen and rng.max_row == fila_origen
    ]
    for rng in origen_merges:
        ref = (
            f"{get_column_letter(rng.min_col)}{fila_destino}:"
            f"{get_column_letter(rng.max_col)}{fila_destino}"
        )
        try:
            ws.merge_cells(ref)
        except ValueError:
            pass


def _escribir_celda(ws, referencia: str, valor):
    """La plantilla real tiene L33:M33 fusionada, pero MAPEO_DE_CELDAS del
    GAS original apunta 'profundidad_mm' a L33 (ancla) y
    'reporte_anexo_grafico_no' a M33 (dentro de la misma fusión) — Apps
    Script tolera escribir en una celda no-ancla de un merge sin error
    (queda invisible detrás del merge); openpyxl la marca de solo lectura
    y lanza AttributeError. Se replica el efecto visual real (silenciosa,
    no se pierde el valor de la celda ancla) en vez de fallar."""
    celda = ws[referencia]
    if celda.__class__.__name__ == "MergedCell":
        logger.warning("Celda %s es parte de un merge no-ancla, se omite (igual que en Apps Script)", referencia)
        return
    celda.value = valor


def _copiar_estilo_fila(ws, fila_origen: int, fila_destino: int, max_col: int = 26):
    for c in range(1, max_col + 1):
        origen = ws.cell(row=fila_origen, column=c)
        destino = ws.cell(row=fila_destino, column=c)
        destino.font = copy(origen.font)
        destino.border = copy(origen.border)
        destino.fill = copy(origen.fill)
        destino.alignment = copy(origen.alignment)
        destino.number_format = origen.number_format


def generar_reporte_acfm(
    fila_general: dict,
    secciones_data: dict[str, list[dict]],
    secciones_fotos: dict[str, list[dict]],
    progreso=None,
) -> bytes:
    """Genera el .xlsx real de Insp ACFM y devuelve los bytes.

    `secciones_data["datosACFM"]` = filas de '1.1_reporte_datos' filtradas
    por id_general. `secciones_data["fotosGenerales"]` siempre vacío (no
    tiene hoja de datos propia). `secciones_fotos[key]` = fotos filtradas
    por id_general en ambos casos.
    """
    def _reportar(pct: int, etapa: str):
        if progreso:
            progreso(pct, etapa)

    _reportar(3, "Preparando plantilla")
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[HOJA_FORMATO]
    desactivar_fit_to_page(ws)

    _reportar(5, "Escribiendo datos generales")
    for campo, celda in CELDAS_GENERALES.items():
        valor = fila_general.get(campo)
        if valor:
            ws[celda] = valor_tipado(valor)

    _reportar(7, "Insertando imágenes de esquema/registro/firma")
    esquema_bytes = descargar_imagen(fila_general.get("link_imagen_esquema", ""))
    if esquema_bytes:
        insertar_imagen_centrada(ws, esquema_bytes, CELDA_ESQUEMA)
    registro_bytes = descargar_imagen(fila_general.get("link_imagen_registro", ""))
    if registro_bytes:
        insertar_imagen_centrada(ws, registro_bytes, CELDA_REGISTRO)
    firma_bytes = descargar_imagen(fila_general.get("link_firma", ""))
    if firma_bytes:
        insertar_imagen_centrada(ws, firma_bytes, CELDA_FIRMA)

    total_fotos = sum(len(f) for f in secciones_fotos.values()) or 1
    fotos_procesadas = 0
    filas_acumuladas = 0

    for idx_sec, key in enumerate(SECTION_KEYS_ORDEN):
        config = SECTIONS_CONFIG[key]
        registros = secciones_data.get(key, [])
        fotos = secciones_fotos.get(key, [])

        pct_base = 10 + round((idx_sec / len(SECTION_KEYS_ORDEN)) * 80)
        _reportar(pct_base, f"Sección {idx_sec + 1}/{len(SECTION_KEYS_ORDEN)}")

        fila_inicio = config["data_start_row"] + filas_acumuladas
        # Plantilla ACFM tiene 1 sola fila de capacidad (no 2 como 570/PM) —
        # ver docstring del módulo, punto 1.
        filas_extra_datos = max(0, len(registros) - 1) if config["mapping"] else 0

        if filas_extra_datos > 0:
            _insertar_filas_y_ajustar_alturas(ws, fila_inicio + 1, filas_extra_datos)
            altura_patron = ws.row_dimensions[fila_inicio].height
            for i in range(filas_extra_datos):
                fila_nueva = fila_inicio + 1 + i
                _copiar_estilo_fila(ws, fila_inicio, fila_nueva)
                _replicar_merges_de_fila(ws, fila_inicio, fila_nueva)
                if altura_patron:
                    ws.row_dimensions[fila_nueva].height = altura_patron

        if config["mapping"]:
            for i, reg in enumerate(registros):
                fila_actual = fila_inicio + i
                for campo, col in config["mapping"].items():
                    valor = reg.get(campo)
                    if valor:
                        _escribir_celda(ws, f"{col}{fila_actual}", valor_tipado(valor))

        filas_acumuladas += filas_extra_datos

        # ---- Fotos de esta sección ----
        base_photo_row = config["photo_row"] + filas_acumuladas
        base_desc_row = config["desc_row"] + filas_acumuladas

        chunks_necesarios = -(-max(len(fotos), 1) // 4)  # ceil(n/4), mínimo 1
        if chunks_necesarios > 1:
            filas_extra_fotos = (chunks_necesarios - 1) * 2
            _insertar_filas_y_ajustar_alturas(ws, base_photo_row + 1, filas_extra_fotos)
            for c in range(1, chunks_necesarios):
                f_foto = base_photo_row + c * 2
                f_desc = f_foto + 1
                _copiar_estilo_fila(ws, base_photo_row, f_foto)
                _copiar_estilo_fila(ws, base_desc_row, f_desc)
                _replicar_merges_de_fila(ws, base_photo_row, f_foto)
                _replicar_merges_de_fila(ws, base_desc_row, f_desc)
                ws.row_dimensions[f_foto].height = ws.row_dimensions[base_photo_row].height
                ws.row_dimensions[f_desc].height = ws.row_dimensions[base_desc_row].height
            filas_acumuladas += filas_extra_fotos

        for i, foto in enumerate(fotos):
            chunk_idx = i // 4
            pos_in_chunk = i % 4
            f_foto = base_photo_row + chunk_idx * 2
            f_desc = f_foto + 1
            col = config["photo_cols"][pos_in_chunk]

            desc = foto.get("descripcion") or ""
            if desc:
                ws[f"{col}{f_desc}"] = desc

            img_bytes = descargar_imagen(foto.get("url") or "")
            if img_bytes:
                insertar_imagen_centrada(ws, img_bytes, f"{col}{f_foto}")

            fotos_procesadas += 1
            pct_fotos = 10 + round((fotos_procesadas / total_fotos) * 80)
            _reportar(min(pct_fotos, 90), f"Foto {fotos_procesadas} de {total_fotos}")

    _reportar(97, "Guardando archivo")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
