"""
Generador REAL de reportes API 510 (Inspección Visual de Recipientes a
Presión) usando la plantilla verificada (templates_xlsx/510.xlsx) y datos
reales de DOS Google Sheets separados: `DB_INSP_Visual_510_ADC` (datos) y
`PHOTOS_VT510_APPSHEETS` (fotos) — a diferencia de 570, donde todo vive en
un único Sheet. Mismo patrón que report_engine_570.py, generalizado a 11
secciones (ver APP003_Recipie_510_VT/APP003_Recipie_510_VT.js — SECTIONS_CONFIG).

Diferencia estructural frente a 570: la plantilla de 510 solo trae UNA fila
de capacidad por sección (no dos) — verificado celda por celda contra
templates_xlsx/510.xlsx el 2026-07-03.
"""
import io
import logging
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .image_utils import desactivar_fit_to_page, descargar_imagen, insertar_imagen_centrada

logger = logging.getLogger("report_engine_510")

TEMPLATE_PATH = Path(__file__).resolve().parent / "templates_xlsx" / "510.xlsx"
HOJA_FORMATO = "FORMATO_VISUAL"

# Verificado contra los headers reales de '0.pv_general' (2026-07-03) — los
# nombres de campo son distintos a los del script GAS original en algunos
# casos porque la hoja tiene columnas truncadas/renombradas desde entonces
# (ej. 'matChaqueta' en el GAS ya no existe, la columna real es 'matchaque').
CELDAS_GENERALES = {
    "cliente": "I8", "consecutivo": "V8", "fechainsp": "AI8", "ubicación": "AV8",
    "tag": "I10", "servicio": "V10", "fabricante": "AI10", "yearfabrication": "AV10",
    "nbno": "I12", "noserie": "V12", "mawp": "AI12", "designtemp": "AV12",
    "rt": "I14", "mdmt": "V14", "po": "AI14", "opertemp": "AV14",
    "ca": "I16", "code": "V16", "alturalargo": "AI16", "diametro": "AV16",
    "matcuerpo": "I18", "matcabezas": "V18", "mattapa": "AI18", "matchaque": "AV18",
    "thkcuerpo": "I20", "thkcabeza": "V20", "thktapa": "AI20", "capacidad": "AV20",
}
CELDA_FOTO_GENERAL_1 = "C23"
CELDA_FOTO_GENERAL_2 = "AC23"
CELDA_FIRMA = "J125"
CELDA_NOMBRE_FIRMA = "J126"

# Traducido literal de SECTIONS_CONFIG en APP003_Recipie_510_VT.js (2026-07-03)
SECTIONS_CONFIG = {
    "foundation": {
        "sheet": "1.pv_foundation", "photo_sheet": "1.pv_foundation_photos", "photo_link_col": "link_pv_foundation",
        "data_start_row": 28, "mapping": {"fundconcreto1": "C", "fundconcreto2": "I", "fundconcreto3": "O", "fundconcreto4": "T", "conten1": "AC", "conten2": "AI", "conten3": "AO", "conten4": "AU"},
        "photo_row": 31, "desc_row": 32, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "support": {
        "sheet": "2.pv_support", "photo_sheet": "2.pv_support_photos", "photo_link_col": "link_pv_support",
        "data_start_row": 37, "mapping": {"support1": "C", "support2": "H", "support3": "N", "support4": "T", "support5": "Z", "support6": "AF", "support7": "AL"},
        "photo_row": 40, "desc_row": 41, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "shellExternal": {
        "sheet": "3.pv_shell_External", "photo_sheet": "3.pv_shell_External_photos", "photo_link_col": "link_pv_shell_External",
        "data_start_row": 46, "mapping": {"pvexternal1": "C", "pvexternal2": "G", "pvexternal3": "L", "pvexternal4": "Q", "pvexternal5": "V", "pvexternal6": "AA", "pvexternal7": "AF", "pvexternal8": "AK", "pvexternal9": "AN", "pvexternal10": "AQ"},
        "photo_row": 49, "desc_row": 50, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "shellInternal": {
        "sheet": "4.pv_shell_Internal", "photo_sheet": "4.pv_shell_Internal_photos", "photo_link_col": "link_pv_shell_Internal",
        "data_start_row": 55, "mapping": {"pvinternal1": "C", "pvinternal2": "G", "pvinternal3": "L", "pvinternal4": "Q", "pvinternal5": "V", "pvinternal6": "AA", "pvinternal7": "AF", "pvinternal8": "AK"},
        "photo_row": 58, "desc_row": 59, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "nozzle": {
        "sheet": "5.pv_nozzle", "photo_sheet": "5.pv_nozzle_photos", "photo_link_col": "link_pv_nozzle",
        "data_start_row": 64, "mapping": {"pvnozzle1": "C", "pvnozzle2": "E", "pvnozzle3": "H", "pvnozzle4": "J", "pvnozzle5": "O", "pvnozzle6": "R", "pvnozzle7": "V", "pvnozzle8": "Z", "pvnozzle9": "AD", "pvnozzle10": "AH", "pvnozzle11": "AL", "pvnozzle12": "AP", "pvnozzle13": "AT"},
        "photo_row": 67, "desc_row": 68, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "prd": {
        "sheet": "6.pv_PRD", "photo_sheet": "6.pv_PRD_photos", "photo_link_col": "link_pv_PRD",
        "data_start_row": 73, "mapping": {"prd1": "C", "prd2": "G", "prd3": "K", "prd4": "N", "prd5": "R", "prd6": "U", "prd7": "X", "prd8": "AA", "prd9": "AD", "prd10": "AH", "prd11": "AL", "prd13": "AP", "prd14": "AU"},
        "photo_row": 76, "desc_row": 77, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "ladders": {
        "sheet": "7.pv_Ladders_Stairways_Platform", "photo_sheet": "7.pv_Ladders_Stairways_Platform_photos", "photo_link_col": "link_pv_Ladders_Stairways_Platform",
        "data_start_row": 82, "mapping": {"stairways1": "C", "stairways2": "N", "stairways3": "Y", "stairways4": "AJ"},
        "photo_row": 85, "desc_row": 86, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "indicators": {
        "sheet": "8.pv_pressu_temp_indicators", "photo_sheet": "8.pv_pressu_temp_indicators_photos", "photo_link_col": "link_pv_pressu_temp_indicators",
        "data_start_row": 91, "mapping": {"inst1": "C", "inst2": "N", "inst3": "Y", "inst4": "AJ"},
        "photo_row": 94, "desc_row": 95, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "insulation": {
        "sheet": "9.pv_insulation", "photo_sheet": "9.pv_insulation_photos", "photo_link_col": "link_pv_insulation",
        "data_start_row": 100, "mapping": {"aislami1": "C", "aislami2": "N", "aislami3": "Y", "aislami4": "AJ"},
        "photo_row": 103, "desc_row": 104, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "mixer": {
        "sheet": "10.pv_Mixer_Agitator", "photo_sheet": "10.pv_Mixer_Agitator_photos", "photo_link_col": "link_pv_Mixer_Agitator",
        "data_start_row": 109, "mapping": {"mixer1": "C", "mixer2": "N", "mixer3": "Y", "mixer4": "AJ"},
        "photo_row": 112, "desc_row": 113, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "jacket": {
        "sheet": "11.pv_jacket", "photo_sheet": "11.pv_jacket_photos", "photo_link_col": "link_pv_jacket",
        "data_start_row": 118, "mapping": {"jacket1": "C", "jacket2": "K", "jacket3": "S", "jacket4": "AA", "jacket5": "AI", "jacket6": "AQ"},
        "photo_row": 121, "desc_row": 122, "photo_cols": ["C", "P", "AC", "AP"],
    },
}

SECTION_KEYS_ORDEN = list(SECTIONS_CONFIG.keys())


def _insertar_filas_y_ajustar_alturas(ws, pos: int, n: int):
    if n <= 0:
        return
    max_row_antes = ws.max_row
    alturas_originales = {
        r: ws.row_dimensions[r].height
        for r in range(pos, max_row_antes + 1)
        if r in ws.row_dimensions and ws.row_dimensions[r].height is not None
    }
    merges_a_mover = [r for r in list(ws.merged_cells.ranges) if r.min_row >= pos]
    for rng in merges_a_mover:
        ws.unmerge_cells(str(rng))

    ws.insert_rows(pos, n)

    for rng in merges_a_mover:
        rng.shift(0, n)
        ws.merge_cells(str(rng))

    for r in range(pos, max_row_antes + n + 1):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    for r_original, altura in alturas_originales.items():
        ws.row_dimensions[r_original + n].height = altura


def _replicar_merges_de_fila(ws, fila_origen: int, fila_destino: int):
    origen_merges = [
        rng for rng in list(ws.merged_cells.ranges)
        if rng.min_row == fila_origen and rng.max_row == fila_origen
    ]
    for rng in origen_merges:
        ref = (
            f"{get_column_letter(rng.min_col)}{fila_destino}:"
            f"{get_column_letter(rng.max_col)}{fila_destino}"
        )
        try:
            ws.merge_cells(ref)
        except ValueError:
            pass


def _copiar_estilo_fila(ws, fila_origen: int, fila_destino: int, max_col: int = 56):
    for c in range(1, max_col + 1):
        origen = ws.cell(row=fila_origen, column=c)
        destino = ws.cell(row=fila_destino, column=c)
        destino.font = copy(origen.font)
        destino.border = copy(origen.border)
        destino.fill = copy(origen.fill)
        destino.alignment = copy(origen.alignment)
        destino.number_format = origen.number_format


def generar_reporte_510(
    fila_general: dict,
    secciones_data: dict[str, list[dict]],
    secciones_fotos: dict[str, list[dict]],
    progreso=None,
) -> bytes:
    """Genera el .xlsx real de API 510 y devuelve los bytes.

    A diferencia de 570, la plantilla solo trae UNA fila de capacidad por
    sección (no dos) — `filas_extra_datos = len(registros) - 1`.
    """
    def _reportar(pct: int, etapa: str):
        if progreso:
            progreso(pct, etapa)

    _reportar(3, "Preparando plantilla")
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[HOJA_FORMATO]
    desactivar_fit_to_page(ws)

    _reportar(5, "Escribiendo datos generales")
    for campo, celda in CELDAS_GENERALES.items():
        valor = fila_general.get(campo)
        if valor:
            ws[celda] = valor

    foto1 = descargar_imagen(fila_general.get("photos_link", ""))
    if foto1:
        insertar_imagen_centrada(ws, foto1, CELDA_FOTO_GENERAL_1)
    foto2 = descargar_imagen(fila_general.get("photos_2_link", ""))
    if foto2:
        insertar_imagen_centrada(ws, foto2, CELDA_FOTO_GENERAL_2)

    total_fotos = sum(len(f) for f in secciones_fotos.values()) or 1
    fotos_procesadas = 0
    filas_acumuladas = 0

    for idx_sec, key in enumerate(SECTION_KEYS_ORDEN):
        config = SECTIONS_CONFIG[key]
        registros = secciones_data.get(key, [])
        fotos = secciones_fotos.get(key, [])

        pct_base = 10 + round((idx_sec / len(SECTION_KEYS_ORDEN)) * 80)
        _reportar(pct_base, f"Sección {idx_sec + 1}/{len(SECTION_KEYS_ORDEN)}: {config['sheet']}")

        fila_inicio = config["data_start_row"] + filas_acumuladas
        # La plantilla de 510 solo trae 1 fila de capacidad (no 2 como en 570)
        filas_extra_datos = max(0, len(registros) - 1)

        if filas_extra_datos > 0:
            _insertar_filas_y_ajustar_alturas(ws, fila_inicio + 1, filas_extra_datos)
            altura_patron = ws.row_dimensions[fila_inicio].height
            for i in range(filas_extra_datos):
                fila_nueva = fila_inicio + 1 + i
                _copiar_estilo_fila(ws, fila_inicio, fila_nueva)
                _replicar_merges_de_fila(ws, fila_inicio, fila_nueva)
                if altura_patron:
                    ws.row_dimensions[fila_nueva].height = altura_patron

        for i, reg in enumerate(registros):
            fila_actual = fila_inicio + i
            for campo, col in config["mapping"].items():
                valor = reg.get(campo)
                if valor:
                    ws[f"{col}{fila_actual}"] = valor

        filas_acumuladas += filas_extra_datos

        # ---- Fotos de esta sección ----
        base_photo_row = config["photo_row"] + filas_acumuladas
        base_desc_row = config["desc_row"] + filas_acumuladas

        chunks_necesarios = -(-max(len(fotos), 1) // 4)  # ceil(n/4), mínimo 1
        if chunks_necesarios > 1:
            filas_extra_fotos = (chunks_necesarios - 1) * 2
            _insertar_filas_y_ajustar_alturas(ws, base_photo_row + 1, filas_extra_fotos)
            for c in range(1, chunks_necesarios):
                f_foto = base_photo_row + c * 2
                f_desc = f_foto + 1
                _copiar_estilo_fila(ws, base_photo_row, f_foto)
                _copiar_estilo_fila(ws, base_desc_row, f_desc)
                _replicar_merges_de_fila(ws, base_photo_row, f_foto)
                _replicar_merges_de_fila(ws, base_desc_row, f_desc)
                ws.row_dimensions[f_foto].height = ws.row_dimensions[base_photo_row].height
                ws.row_dimensions[f_desc].height = ws.row_dimensions[base_desc_row].height
            filas_acumuladas += filas_extra_fotos

        for i, foto in enumerate(fotos):
            chunk_idx = i // 4
            pos_in_chunk = i % 4
            f_foto = base_photo_row + chunk_idx * 2
            f_desc = f_foto + 1
            col = config["photo_cols"][pos_in_chunk]

            desc = foto.get("descripcion") or ""
            if desc:
                ws[f"{col}{f_desc}"] = desc

            img_bytes = descargar_imagen(foto.get("url") or "")
            if img_bytes:
                insertar_imagen_centrada(ws, img_bytes, f"{col}{f_foto}")

            fotos_procesadas += 1
            pct_fotos = 10 + round((fotos_procesadas / total_fotos) * 80)
            _reportar(min(pct_fotos, 90), f"Foto {fotos_procesadas} de {total_fotos}")

    # La firma va DESPUÉS de las 11 secciones — su fila real se corre hacia
    # abajo por cada fila que se haya insertado dinámicamente arriba (bug
    # encontrado el 2026-07-03: escribirla en una fila fija producía
    # "MergedCell... read-only" en cuanto alguna sección insertaba filas).
    _reportar(93, "Insertando firma")
    fila_firma = int(CELDA_FIRMA[1:]) + filas_acumuladas
    fila_nombre = int(CELDA_NOMBRE_FIRMA[1:]) + filas_acumuladas
    firma_bytes = descargar_imagen(fila_general.get("link_firma", ""))
    if firma_bytes:
        insertar_imagen_centrada(ws, firma_bytes, f"J{fila_firma}")
    if fila_general.get("nombre"):
        ws[f"J{fila_nombre}"] = fila_general.get("nombre")

    _reportar(97, "Guardando archivo")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
