"""
Generador REAL de reportes PMI — Caracterización de Materiales (.xlsx) usando
la plantilla verificada (templates_xlsx/PMI.xlsx) y datos reales del Sheet.
Traducción directa de ADEMINCOL-Scripts/APP004_Caract_Mat_PMI/APP004_Caract_Mat_PMI.js,
verificada celda por celda contra la plantilla real el 2026-07-03.

A diferencia de MT, la plantilla PMI usa RANGOS FIJOS pre-formateados para
química (18 slots) y durezas (59 slots) — no hace falta insertar filas
dinámicamente, lo que simplifica bastante el motor.
"""
import io
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string

from .chart_durezas import ELEMENTO_DEFAULT, generar_grafico_durezas
from .image_utils import desactivar_fit_to_page, descargar_imagen, insertar_imagen_centrada
from .report_utils import valor_tipado

logger = logging.getLogger("report_engine_pmi")

TEMPLATE_PATH = Path(__file__).resolve().parent / "templates_xlsx" / "PMI.xlsx"

# Verificado celda por celda contra la plantilla real y contra
# APP004_Caract_Mat_PMI.js (MAPEO_GENERAL) el 2026-07-03.
CELDAS_GENERALES = {
    "Cliente": "E7", "Contrato": "M7", "N_Reporte": "R7", "OT": "Z7", "Fecha": "AE7",
    "Departamento": "E9", "Ciudad": "M9", "Troncal": "U9", "Estación": "AC9",
    "Sistema": "E11", "Linea": "R11", "PK": "AC11",
    "Equipo_Inspeccionado": "G16", "Tag": "N16",
    "Descripcion_Componente": "G18", "Estado_Componente": "G20", "Observacion_Estado": "M20",
    "Ubicacion_Componente": "G22", "Dimensiones": "G24",
    "NPS": "G26", "Espesor_Min_Pulg": "M26",
    "Plano_Referencia": "G28", "Observaciones_Generales": "B31",
    # Metalografía (1_M)
    "1_M_Procedimiento": "G54", "1_M_Tecnica": "Q54", "1_M_Normas_Referencia": "AC54",
    "1_M_Abrasivo": "F54",
    "1_M_Equipo_Desbaste": "F60", "1_M_Marca_Desbaste": "M60",
    "1_M_Modelo_Desbaste": "F62", "1_M_Serie_Desbaste": "M62",
    "1_M_Micro_Marca": "V60", "1_M_Micro_Modelo": "AC60",
    "1_M_Micro_Serie": "V62", "1_M_Micro_Lentes": "AC62",
    "1_M_Material_Analizar": "J68", "1_M_Tiempo_Ataque_Seg": "J69",
    "1_M_Reactivo_Norma": "AB68",
    "1_M_Calc_Vol_Solucion": "AB73", "1_M_Calc_Conc_Acido_Base": "AB74",
    "1_M_Calc_Conc_Deseada": "AB75", "1_M_Res_Vol_Acido": "AB75", "1_M_Res_Vol_Dilusor": "AB76",
    "1_M_aumentos_metalografias": "L79",
    "1_M_comentario_2": "P84", "1_M_comentario_3": "Z84",
    "1_M_analisis_inclusiones": "L86",
    "1_M_comentario_4": "P91", "1_M_comentario_5": "Z91",
    "1_M_analisis_de_inclusiones": "L93",
    "1_M_comentario_6": "U99",
    "1_M_tamano_grano": "F101", "1_M_fases": "N101",
    "1_M_porceso_fabricacion": "V101", "1_M_defectos": "AD101",
    "1_M_analisis_metalografico": "L103",
    "Material_referencia": "B125", "Material_referencia_2": "B130",
    # Química (2_Q)
    "2_Q_Procedimiento": "G110", "2_Q_Tecnica": "Q110", "2_Q_Normas_Referencia": "AC110",
    "2_Q_Equipo_Desbaste": "F114", "2_Q_Marca_Desbaste": "V114",
    "2_Q_Modelo_Desbaste": "F116", "2_Q_Serie_Desbaste": "Q116",
    "2_Q_fecha_calibracion": "AB116",
    "2_Q_comentario_7": "E163", "2_Q_comentario_8": "U163",
    # Dureza (3_D)
    "3_D_Procedimiento": "G169", "3_D_Tecnica": "Q169", "3_D_Normas_Referencia": "AC169",
    "3_D_Marca_Durometro": "F173", "3_D_Modelo_Durometro": "F175",
    "3_D_Serie_Durometro": "W173", "3_D_Fecha_Calibracion": "W175",
    "3_D_Ubicacion_Horaria": "E179", "3_D_Escala_Dureza": "M179",
    "3_D_Tolerancia": "U179", "3_D_Material_Referencia": "AC179",
    "3_D_comentario_9": "E219", "3_D_comentario_10": "U219",
    "3_D_analisis_mecanicas": "B221",
    # Extras
    "comentario_1": "U32", "nombre": "G224", "cargo": "G225",
}

CELDAS_IMAGENES = {
    "link_foto": "R16", "link_imagen_2": "M83", "link_imagen_3": "W83",
    "link_imagen_4": "M90", "link_imagen_5": "W90", "link_imagen_6": "R98",
    "link_imagen_7": "B146", "link_imagen_8": "R146",
    "link_imagen_9": "B202", "link_imagen_10": "R202",
    "link_firma": "G223",
}

# Bloque de firmas (fila 222: "REALIZADO POR" / "REVISADO POR" / "APROBADO
# POR"). G223-226 (realizado = inspector) ya se llenaba con 'link_firma'/
# 'nombre'/'cargo'. P223-226 (revisado = supervisor que genera el reporte
# desde la webapp) es nuevo (decisión 2026-07-05, primera prueba con PMI):
# se llena con el usuario autenticado en la plataforma, NO con un dato del
# Sheet de PMI — main.py lo resuelve contra la BD de usuarios y lo pasa aquí
# como 'supervisor_nombre'/'supervisor_cargo'/'supervisor_firma_link'.
CELDA_FIRMA_SUPERVISOR = "P223"
CELDA_NOMBRE_SUPERVISOR = "P224"
CELDA_CARGO_SUPERVISOR = "P225"
CELDA_FECHA_SUPERVISOR = "P226"

# B221 (celda combinada B221:AG221, vacía en la plantilla) lista los puntos
# atípicos de durezas que se quitaron del gráfico automático (decisión
# 2026-07-08, ver generar_grafico_durezas en chart_durezas.py).
CELDA_ATIPICOS_DUREZAS = "B221"

# Rangos fijos de la tabla de química (18 slots: 6 filas x 3 columnas).
# NOTA: la plantilla real tiene 7 filas por columna (21 slots, ver fila 142
# con índice "7"/"14"/"21"), pero el script GAS original solo usa 6
# (136:141) — confirmado contra la plantilla el 2026-07-03. Si algún informe
# trae más de 18 elementos químicos analizados, los adicionales se pierden
# silenciosamente (igual que en el GAS actual). Documentado, no corregido
# aquí para no desviarse del comportamiento ya validado en producción.
RANGO_QUIMICA_ELEMENTO = ["D136", "D137", "D138", "D139", "D140", "D141",
                           "N136", "N137", "N138", "N139", "N140", "N141",
                           "X136", "X137", "X138", "X139", "X140", "X141"]
RANGO_QUIMICA_VALOR = ["G136", "G137", "G138", "G139", "G140", "G141",
                        "Q136", "Q137", "Q138", "Q139", "Q140", "Q141",
                        "AA136", "AA137", "AA138", "AA139", "AA140", "AA141"]

# Rangos fijos de la tabla de durezas (59 slots: filas 184-198, 4 columnas;
# la 4ª columna se detiene en 197 porque la fila 198 tiene la fórmula de
# promedio "=IFERROR(AVERAGE(...))" que NO debe sobrescribirse).
def _rango(col: str, fila_ini: int, fila_fin: int) -> list[str]:
    return [f"{col}{f}" for f in range(fila_ini, fila_fin + 1)]

RANGO_DUREZAS = (
    _rango("F", 184, 198) + _rango("M", 184, 198) + _rango("U", 184, 198) + _rango("AB", 184, 197)
)
RANGO_KSI = (
    _rango("H", 184, 198) + _rango("O", 184, 198) + _rango("W", 184, 198) + _rango("AD", 184, 197)
)

ELEM_KEY = {
    "c (carbono)": "c", "c": "c", "mn (manganeso)": "mn", "mn": "mn",
    "si (silicio)": "si", "si": "si", "cu (cobre)": "cu", "cu": "cu",
    "ni (niquel)": "ni", "ni (níquel)": "ni", "ni": "ni",
    "cr (cromo)": "cr", "cr": "cr", "mo (molibdeno)": "mo", "mo": "mo",
    "v (vanadio)": "v", "v": "v", "b (boro)": "b", "b": "b",
}


def extraer_ksis(durezas: list[dict]) -> list[float]:
    """Valores numéricos de ksi en el orden en que vienen las durezas —
    usado tanto al generar el reporte como en el endpoint de previsualización
    del gráfico (ver /api/preview/pmi/{id}/grafico-durezas en main.py)."""
    ksis = []
    for row in durezas:
        raw = str(row.get("ksi", "")).strip().replace(",", ".")
        try:
            ksis.append(float(raw))
        except ValueError:
            continue
    return ksis


def calcular_ce(quimica: list[dict]) -> float | None:
    """Carbono Equivalente — misma fórmula que calcularCE() en el GAS
    (CE_Pcm si C<=0.12%, CE_IIW si C>0.12%). Se recalcula aquí porque el
    trigger de Sheets solo corre cuando AppSheet escribe, no bajo demanda."""
    valores: dict[str, list[float]] = {}
    for row in quimica:
        elem = ELEM_KEY.get(str(row.get("Elemento", "")).strip().lower())
        if not elem:
            continue
        raw = str(row.get("Valor", "")).strip().replace("%", "").replace(",", ".")
        try:
            val = float(raw)
        except (ValueError, TypeError):
            continue
        valores.setdefault(elem, []).append(val)

    def prom(key: str) -> float:
        vals = valores.get(key, [])
        return sum(vals) / len(vals) if vals else 0.0

    c, mn, si, cu = prom("c"), prom("mn"), prom("si"), prom("cu")
    ni, cr, mo, v, b = prom("ni"), prom("cr"), prom("mo"), prom("v"), prom("b")
    if not valores:
        return None
    if c <= 0.12:
        ce = c + si / 30 + mn / 20 + cu / 20 + ni / 60 + cr / 20 + mo / 15 + v / 10 + 5 * b
    else:
        ce = c + mn / 6 + (cr + mo + v) / 5 + (ni + cu) / 15
    return round(ce, 4)


def generar_reporte_pmi(
    fila_general: dict,
    quimica: list[dict],
    durezas: list[dict],
    progreso=None,
) -> bytes:
    """Genera el .xlsx real de PMI y devuelve los bytes del archivo."""
    def _reportar(pct: int, etapa: str):
        if progreso:
            progreso(pct, etapa)

    _reportar(5, "Preparando plantilla")
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["FORMATO_MATERIALES"]
    desactivar_fit_to_page(ws)

    _reportar(15, "Escribiendo datos generales")
    # CELDAS_GENERALES conserva las mayúsculas del script GAS original (para
    # comparar fácil contra la fuente), pero `read_sheet_as_dicts` normaliza
    # los headers del Sheet a minúsculas — se hace match case-insensitive.
    fila_general_lower = {str(k).lower(): v for k, v in fila_general.items()}
    for campo, celda in CELDAS_GENERALES.items():
        valor = fila_general_lower.get(campo.lower())
        if valor in (None, ""):
            continue
        destino = ws[celda]
        if type(destino).__name__ == "MergedCell":
            # Dos campos del mapeo original (1_M_Abrasivo->F54,
            # 1_M_Res_Vol_Dilusor->AB76) apuntan a una celda que NO es la
            # esquina superior-izquierda de su combinación — confirmado
            # contra la plantilla real el 2026-07-03. openpyxl no permite
            # escribir ahí (a diferencia de Sheets, que lo tolera en
            # silencio). Se omiten con warning en vez de fallar todo el
            # reporte por dos campos de metadata no críticos.
            logger.warning(
                "Campo '%s' apunta a %s, que es una celda combinada no-ancla; se omite.",
                campo, celda,
            )
            continue
        ws[celda] = valor_tipado(valor)

    _reportar(30, "Calculando química y durezas")
    # Química: promediar mediciones repetidas por elemento (igual al GAS),
    # y llenar los slots fijos en orden.
    acumulado: dict[str, dict] = {}
    for row in quimica:
        elem = str(row.get("Elemento", "")).strip()
        raw = str(row.get("Valor", "")).strip().replace("%", "").replace(",", ".")
        if not elem or not raw:
            continue
        try:
            val = float(raw)
        except ValueError:
            continue
        acumulado.setdefault(elem, {"suma": 0.0, "n": 0})
        acumulado[elem]["suma"] += val
        acumulado[elem]["n"] += 1

    for idx, (elem, datos) in enumerate(acumulado.items()):
        if idx >= len(RANGO_QUIMICA_ELEMENTO):
            logger.warning("Más de %d elementos químicos, se truncan (igual que el GAS actual)",
                            len(RANGO_QUIMICA_ELEMENTO))
            break
        ws[RANGO_QUIMICA_ELEMENTO[idx]] = elem
        # Bug encontrado 2026-07-09: 'Valor' en el Sheet trae el % ya escrito
        # (ej. "1.12%" -> tras el strip() de arriba queda 1.12, el NÚMERO
        # 1.12 no la fracción 0.0112). Las celdas G136/Q136/AA136... tienen
        # number_format '0.00%' en la plantilla — Excel multiplica por 100 al
        # mostrar un valor con formato de porcentaje, así que escribir 1.12
        # tal cual se veía como "112.00%" (y 93.50 -> "9350.00%"). Se divide
        # entre 100 aquí, al ESCRIBIR en Excel, para volver a la fracción que
        # el formato de la celda espera — calcular_ce() sigue usando el
        # promedio SIN dividir (datos["suma"]/datos["n"]), porque la fórmula
        # de Carbono Equivalente ya está definida en términos del número de
        # porcentaje (ej. C=0.12 para 0.12%), no de la fracción.
        ws[RANGO_QUIMICA_VALOR[idx]] = (datos["suma"] / datos["n"]) / 100

    # Durezas: se listan en el orden que vienen, sin promediar (igual al GAS)
    for idx, row in enumerate(durezas):
        if idx >= len(RANGO_DUREZAS):
            logger.warning("Más de %d mediciones de dureza, se truncan (igual que el GAS actual)",
                            len(RANGO_DUREZAS))
            break
        dureza = row.get("Dureza")
        ksi = row.get("ksi")
        if dureza not in (None, ""):
            ws[RANGO_DUREZAS[idx]] = valor_tipado(dureza)
        if ksi not in (None, "") and idx < len(RANGO_KSI):
            ws[RANGO_KSI[idx]] = valor_tipado(ksi)

    _reportar(45, "Insertando imágenes")
    total_imgs = len(CELDAS_IMAGENES)
    for i, (campo, celda) in enumerate(CELDAS_IMAGENES.items()):
        url = fila_general.get(campo)
        pct = 45 + round((i / max(total_imgs, 1)) * 50)

        # link_imagen_10 (celda R202) es el gráfico Tensión vs Punto que
        # antes se generaba corriendo un script en R y se subía a mano.
        # Decisión 2026-07-08: el gráfico SIEMPRE se genera aquí, incluso si
        # el inspector subió una imagen manual — deja de respetarse esa
        # imagen (antes se usaba como prioridad; ahora el gráfico automático
        # la reemplaza en todos los casos).
        if campo == "link_imagen_10":
            _reportar(pct, "Generando gráfico de durezas")
            elemento = str(fila_general.get("elemento_grafico") or ELEMENTO_DEFAULT).strip().upper()
            grafico_bytes, resumen_atipicos = generar_grafico_durezas(extraer_ksis(durezas), elemento)
            if grafico_bytes:
                col, fila = coordinate_from_string(celda)
                insertar_imagen_centrada(ws, grafico_bytes, f"{col}{fila}")
            # Puntos atípicos (fuera de [Lim. Inf., Lim. Sup.]) que se
            # quitaron del gráfico — se listan aquí en vez de dibujarse.
            if resumen_atipicos:
                ws[CELDA_ATIPICOS_DUREZAS] = resumen_atipicos
            continue

        if not url:
            continue
        _reportar(pct, f"Descargando imagen {i + 1} de {total_imgs}")
        img_bytes = descargar_imagen(url)
        if img_bytes:
            col, fila = coordinate_from_string(celda)
            insertar_imagen_centrada(ws, img_bytes, f"{col}{fila}")

    # Bloque "REVISADO POR" (P223-226) — supervisor que generó el reporte
    # desde la webapp (ver comentario en CELDA_NOMBRE_SUPERVISOR arriba).
    supervisor_nombre = fila_general.get("supervisor_nombre")
    if supervisor_nombre:
        _reportar(96, "Escribiendo datos del supervisor")
        ws[CELDA_NOMBRE_SUPERVISOR] = supervisor_nombre
        if fila_general.get("supervisor_cargo"):
            ws[CELDA_CARGO_SUPERVISOR] = fila_general["supervisor_cargo"]
        ws[CELDA_FECHA_SUPERVISOR] = datetime.now().strftime("%Y-%m-%d")
        firma_bytes = descargar_imagen(fila_general.get("supervisor_firma_link", ""))
        if firma_bytes:
            insertar_imagen_centrada(ws, firma_bytes, CELDA_FIRMA_SUPERVISOR)

    _reportar(97, "Guardando archivo")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
