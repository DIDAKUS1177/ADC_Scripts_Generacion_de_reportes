"""
Generador REAL de reportes MT (.xlsx) usando la plantilla verificada
(templates_xlsx/MT.xlsx) y datos reales del Sheet. Es una versión adelantada
y simplificada del motor genérico de la Fase 4 (ver docs/04_GENERACION_REPORTES.md)
— aquí está hardcodeado para MT porque es solo para probar el flujo completo
antes de tener PostgreSQL. El motor genérico (config por tipo, sin ifs) se
construye en la Fase 4.
"""
import io
import logging
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string

from .image_utils import descargar_imagen, insertar_imagen_centrada
from .report_utils import valor_tipado

logger = logging.getLogger("report_engine_mt")

TEMPLATE_PATH = Path(__file__).resolve().parent / "templates_xlsx" / "MT.xlsx"

# Verificado celda por celda contra la plantilla real (ver docs/04_GENERACION_REPORTES.md)
CELDAS_GENERALES = {
    "cliente": "C7", "contrato": "H7", "ot": "K7", "fecha_actividad": "N7", "reporte_n": "R7",
    "zona": "C9", "sistema": "I9", "subsistema_linea": "O9",
    "departamento": "C11", "municipio": "I11", "pk_sistema": "O11", "distancia_registro": "S11",
    "descripcion_elemento": "F15", "acabado_superficial": "R15",
    "material": "D17", "espesor": "J17", "diametro": "N17", "cantidad_inspeccionada": "S17",
    "plano_referencia": "D19",
    "procedimiento_n": "E23", "revision": "K23", "norma_codigo_ref": "Q23",
    "tecnica_magnetizacion": "E25", "fuerza_campo": "L25", "direccion_campo": "S25",
    "tecnica_desmagnetizacion": "F27",
    "tipo_particulas": "D31", "metodo_aplicacion": "K31", "color_particulas": "O31", "tipo_luz_negra": "T31",
    "marca_equipo": "E33", "codigo_equipo": "P33",
    "marca_particulas": "E35", "codigo_particulas": "R35",
    "intensidad_luz_blanca": "E37", "intensidad_luz_negra": "R37",
    "tipo_corriente": "E39", "equipo_medicion_luz": "K39", "equipo_luz_sn": "R39",
    "observaciones": "D52",
    "nombre": "D54",
    "certificado": "D55",
    "fecha": "D57",
}
CELDA_FIRMA = "D56"
FILA_INICIO_INSPECCION = 44
COLUMNAS_RESULTADO = {"item": "A", "identificacion": "B", "evaluacion": "O", "observaciones": "Q"}
COLUMNAS_INDICACIONES = [("I", "J"), ("K", "L"), ("M", "N")]
FILA_BASE_FOTO = 49
FILA_BASE_DESC = 50


def _insertar_filas_y_ajustar_alturas(ws, pos: int, n: int):
    """`ws.insert_rows()` desplaza el CONTENIDO de las celdas pero NO desplaza
    ni `ws.row_dimensions` (alturas de fila) ni `ws.merged_cells.ranges`
    (celdas combinadas) — limitaciones de openpyxl confirmadas empíricamente
    el 2026-07-02 con tests aislados. Consecuencias que esto causaba:
    - La fila alta de fotos (221px) quedaba en el número de fila equivocado
      y las fotos aterrizaban en filas de ~20px.
    - Los merges (A49:K49 del área de foto, B44:D44 de identificación, etc.)
      quedaban huérfanos en sus filas originales; `_insertar_imagen_centrada`
      no encontraba merge en la posición final de la foto y escalaba la
      imagen al ancho de UNA columna (~48px → miniaturas), y la tabla de
      resultados se veía rota (texto desbordado, celdas sin combinar).
    Aquí se desplazan manualmente ambas cosas después de insertar.
    """
    max_row_antes = ws.max_row

    alturas_originales = {
        r: ws.row_dimensions[r].height
        for r in range(pos, max_row_antes + 1)
        if r in ws.row_dimensions and ws.row_dimensions[r].height is not None
    }

    # Merges que empiezan en/debajo del punto de inserción: quitar, desplazar, reponer.
    # (Merges que CRUZAN el punto de inserción no existen en esta plantilla.)
    merges_a_mover = [r for r in list(ws.merged_cells.ranges) if r.min_row >= pos]
    for rng in merges_a_mover:
        ws.unmerge_cells(str(rng))

    ws.insert_rows(pos, n)

    for rng in merges_a_mover:
        rng.shift(0, n)
        ws.merge_cells(str(rng))

    for r in range(pos, max_row_antes + n + 1):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    for r_original, altura in alturas_originales.items():
        ws.row_dimensions[r_original + n].height = altura


def _replicar_merges_de_fila(ws, fila_origen: int, fila_destino: int):
    """Copia los rangos combinados de UNA fila (solo los que empiezan y
    terminan en esa fila) a otra fila. Necesario porque las filas insertadas
    nacen sin merges y la plantilla los usa en la tabla de resultados
    (B:D identificación, E:F zona, G:H diám, O:P evaluación, Q:T observaciones)
    y en las filas de fotos (A:K / L:T)."""
    origen_merges = [
        rng for rng in list(ws.merged_cells.ranges)
        if rng.min_row == fila_origen and rng.max_row == fila_origen
    ]
    for rng in origen_merges:
        ref = (
            f"{get_column_letter(rng.min_col)}{fila_destino}:"
            f"{get_column_letter(rng.max_col)}{fila_destino}"
        )
        try:
            ws.merge_cells(ref)
        except ValueError:
            pass  # ya existe un merge equivalente


def _copiar_estilo_fila(ws, fila_origen: int, fila_destino: int, max_col: int = 20):
    for c in range(1, max_col + 1):
        origen = ws.cell(row=fila_origen, column=c)
        destino = ws.cell(row=fila_destino, column=c)
        destino.font = copy(origen.font)
        destino.border = copy(origen.border)
        destino.fill = copy(origen.fill)
        destino.alignment = copy(origen.alignment)
        destino.number_format = origen.number_format


def generar_reporte_mt(
    fila_general: dict,
    resultados: list[dict],
    indicaciones: list[dict],
    fotos: list[dict],
    progreso=None,
) -> bytes:
    """Genera el .xlsx real y devuelve los bytes del archivo.

    `progreso`: callback opcional `fn(pct: int, etapa: str)` para reportar
    avance (usado por el endpoint asíncrono con barra de progreso).

    IMPORTANTE — orden de operaciones (fuente de un bug ya corregido):
    openpyxl `insert_rows()` desplaza celdas y sus VALORES automáticamente,
    pero NO desplaza imágenes ya insertadas con `add_image()` (ni merges ni
    alturas — eso lo maneja `_insertar_filas_y_ajustar_alturas`). Por eso el
    orden correcto es: 1) calcular e insertar TODAS las filas extra
    (resultados y fotos) primero, 2) escribir todo el texto en las
    posiciones finales, 3) insertar las imágenes al final, cuando ya no
    habrá más inserciones de filas que las desalineen.
    """
    def _reportar(pct: int, etapa: str):
        if progreso:
            progreso(pct, etapa)

    _reportar(5, "Preparando plantilla")
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["FORMATO"]

    # ---- Fase 1: calcular cuántas filas extra hacen falta ----
    filas_extra_resultados = max(0, len(resultados) - 2)  # plantilla ya trae 2 filas (44-45)
    n_fotos = len(fotos)
    pares_fotos_extra = -(-max(0, n_fotos - 2) // 2)  # ceil((n_fotos-2)/2), 0 si n_fotos<=2
    filas_extra_fotos = pares_fotos_extra * 2

    # ---- Fase 2: insertar filas de la tabla de resultados ----
    _reportar(10, "Insertando filas de resultados")
    if filas_extra_resultados > 0:
        fila_patron = FILA_INICIO_INSPECCION + 1  # fila 45, ya con formato
        _insertar_filas_y_ajustar_alturas(ws, fila_patron + 1, filas_extra_resultados)
        altura_patron = ws.row_dimensions[fila_patron].height
        for i in range(filas_extra_resultados):
            fila_nueva = fila_patron + 1 + i
            _copiar_estilo_fila(ws, fila_patron, fila_nueva)
            _replicar_merges_de_fila(ws, fila_patron, fila_nueva)
            if altura_patron:
                ws.row_dimensions[fila_nueva].height = altura_patron

    fila_base_foto_actual = FILA_BASE_FOTO + filas_extra_resultados
    fila_base_desc_actual = FILA_BASE_DESC + filas_extra_resultados

    # ---- Fase 3: insertar filas extra de fotos (pares más allá del primero) ----
    _reportar(15, "Insertando filas de fotos")
    posiciones_fotos: list[tuple[int, int]] = []  # (fila_foto, fila_desc) por índice de foto
    ultima_fila_desc = fila_base_desc_actual
    for p in range(pares_fotos_extra):
        _insertar_filas_y_ajustar_alturas(ws, ultima_fila_desc + 1, 2)
        f_foto, f_desc = ultima_fila_desc + 1, ultima_fila_desc + 2
        _copiar_estilo_fila(ws, fila_base_foto_actual, f_foto)
        _copiar_estilo_fila(ws, fila_base_desc_actual, f_desc)
        _replicar_merges_de_fila(ws, fila_base_foto_actual, f_foto)
        _replicar_merges_de_fila(ws, fila_base_desc_actual, f_desc)
        ws.row_dimensions[f_foto].height = ws.row_dimensions[fila_base_foto_actual].height
        ws.row_dimensions[f_desc].height = ws.row_dimensions[fila_base_desc_actual].height
        posiciones_fotos.append((f_foto, f_desc))
        ultima_fila_desc = f_desc

    def fila_final(fila_original: int) -> int:
        """Todas las inserciones de filas ya ocurrieron: calcula la fila
        final de una celda de la plantilla original según en qué zona cae."""
        if fila_original > FILA_BASE_FOTO:
            return fila_original + filas_extra_resultados + filas_extra_fotos
        if fila_original > FILA_INICIO_INSPECCION:
            return fila_original + filas_extra_resultados
        return fila_original

    # ---- Fase 4: escribir texto (datos generales, resultados, indicaciones, descripciones de fotos) ----
    _reportar(20, "Escribiendo datos generales")
    for campo, celda in CELDAS_GENERALES.items():
        valor = fila_general.get(campo)
        if valor:
            col, fila = coordinate_from_string(celda)
            ws[f"{col}{fila_final(fila)}"] = valor_tipado(valor)

    for idx, res in enumerate(resultados):
        fila_actual = FILA_INICIO_INSPECCION + idx
        for campo, col in COLUMNAS_RESULTADO.items():
            ws[f"{col}{fila_actual}"] = valor_tipado(res.get(campo, ""))
        inds = [i for i in indicaciones if i.get("id_resultado") == res.get("item")]
        for i, (col_tipo, col_long) in enumerate(COLUMNAS_INDICACIONES):
            if i < len(inds):
                ws[f"{col_tipo}{fila_actual}"] = valor_tipado(inds[i].get("tipo", ""))
                ws[f"{col_long}{fila_actual}"] = valor_tipado(inds[i].get("long", ""))

    fila_foto_por_indice: list[tuple[int, int]] = []  # (fila, columna_letra) por foto, para las imágenes
    for idx, foto in enumerate(fotos):
        desc = foto.get("descripcion") or ""
        es_par = idx % 2 == 0
        if idx <= 1:
            f_foto, f_desc = fila_base_foto_actual, fila_base_desc_actual
        else:
            f_foto, f_desc = posiciones_fotos[(idx - 2) // 2]
        col_desc = "B" if es_par else "M"
        ws[f"{col_desc}{f_desc}"] = desc
        fila_foto_por_indice.append((f_foto, "A" if es_par else "L"))

    # ---- Fase 5: insertar imágenes (firma + fotos), ya con todas las filas en su lugar final ----
    _reportar(30, "Insertando firma")
    firma_bytes = descargar_imagen(fila_general.get("firma_link", ""))
    if firma_bytes:
        col_firma, fila_firma = coordinate_from_string(CELDA_FIRMA)
        insertar_imagen_centrada(ws, firma_bytes, f"{col_firma}{fila_final(fila_firma)}")

    # La descarga de fotos es donde va casi todo el tiempo: 35% → 95%
    for idx, foto in enumerate(fotos):
        pct = 35 + round((idx / max(len(fotos), 1)) * 60)
        _reportar(pct, f"Descargando foto {idx + 1} de {len(fotos)}")
        img_bytes = descargar_imagen(foto.get("url") or "")
        if img_bytes:
            f_foto, col_foto = fila_foto_por_indice[idx]
            insertar_imagen_centrada(ws, img_bytes, f"{col_foto}{f_foto}")

    _reportar(97, "Guardando archivo")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
