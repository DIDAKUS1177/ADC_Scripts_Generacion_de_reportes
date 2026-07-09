"""
Generador REAL de reportes SCAN C (Ultrasonido tipo C-Scan), 2026-07-09.
Traducido de los scripts GAS APP013_SCAN_C_Lineas.js (líneas/tubería) y
APP005_C-Scan_RP_AUT.js (recipientes a presión) — ambos comparten la MISMA
estructura de plantilla y motor (SECTIONS_CONFIG con 2 tablas de datos +
1 bloque de fotos), solo cambia el CONFIG (columnas disponibles y celdas
generales difieren un poco entre las dos variantes reales — ver abajo).

NO se hizo "tanques" (FORMAT_tan en la hoja de LINEAS) — pedido explícito
del usuario, solo líneas y recipientes a presión por ahora.

Plantillas: templates_xlsx/SCANC_LINEAS.xlsx y SCANC_RP.xlsx, ambas
extraídas de la hoja real "FORMAT" de cada spreadsheet (exportando con las
credenciales del service account, no hay Excel entregado a mano para este
tipo — igual que 570/510/Espesores en su momento).

Estructura confirmada en la plantilla real (idéntica en ambas variantes):
  - Datos generales: fila 7-15 (ver CELDAS_GENERALES).
  - '4. INFORMACIÓN DE ENSAYO' — dos tablas:
      reporte_datos: filas 21-22 (2 de capacidad base), header en fila 19-20.
      ensayo_datos:  filas 27-28 (2 de capacidad base), header en fila 26.
  - '5. REGISTRO FOTOGRÁFICO': filas 32 (foto) / 33 (descripción),
    4 fotos por bloque (photoCells de 4 columnas), igual patrón que
    report_engine_570.py.
  - Firma + responsable: filas 36-40 (NOMBRE/CARGO/CERTIFICADO/FIRMA/FECHA).
"""
import io
import logging
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .image_utils import desactivar_fit_to_page, descargar_imagen, insertar_imagen_centrada
from .report_utils import valor_tipado

logger = logging.getLogger("report_engine_scanc")

TEMPLATES_DIR = Path(__file__).resolve().parent / "templates_xlsx"
HOJA_FORMATO = "FORMAT"

# --- CELDAS_GENERALES traducido 1:1 de MAPEO_DE_CELDAS en cada script GAS ---
# (2026-07-09, verificado contra la hoja real de cada spreadsheet — no son
# iguales entre las 2 variantes: RP no trae 'seg'/latitud/longitud porque un
# recipiente a presión es un punto fijo, no un tramo con GPS; RP sí trae
# 'fecha_calibracion' que LINEAS no tiene).
SCANC_LINEAS_CONFIG = {
    "spreadsheet_id": "1GXzQAjEK2s0MrM-IundNq2NwCghYZPdhoxA405mHTRg",
    "template": "SCANC_LINEAS.xlsx",
    "hoja_general": "1.0_general",
    "hoja_reporte_datos": "2.0_reporte",
    "hoja_ensayo_datos": "2.1_ensayo",
    "hoja_fotos": "2.0_reporte_photos",
    "celdas_generales": {
        "cliente": "D7", "fecha": "J7", "reporte_n": "Q7", "estacion": "V7",
        "contrato": "D8", "ot": "J8", "zona": "Q8", "sistema": "V8",
        "equipo": "D9", "fluido": "J9", "material": "Q9", "norma": "V9",
        "inicio_inspeccion": "D13", "acoplante": "J13", "estado_superficial": "R13",
        "marca_equipo": "V13", "fin_inspeccion": "D14", "rango_espesores": "J14",
        "temperatura_superficie": "R14", "modelo": "V14", "serie": "D15",
        "tipo_palpador": "J15", "frecuencia": "R15", "tamano": "V15",
        "nombre": "D36", "cargo": "D37", "certificado": "D38", "fecha_firma": "D46",
    },
    "celda_firma": "D39",
    "reporte_datos": {
        "data_start_row": 21,
        "mapping": {
            "id_punto": "A", "sistema_o_linea": "B", "seg": "C", "cml": "D",
            "diametro_in": "E", "tipo_accesorio": "F", "latitud": "G", "longitud": "H",
            "dja_mm": "I", "posicion_horario_inicial": "J", "posicion_horario_final": "K",
            "longitud_barrido_circunferencial_mm": "L", "longitud_barrido_longitudinal_mm": "M",
            "area_barrido": "N", "numero_barridos": "O", "tipo_evaluacion": "P",
            "posicion_horario_soldadura_longitudinal": "Q", "espesor_nominal_mm": "R",
            "espesor_promedio_mm": "S", "espesor_minimo_mm": "T",
            "perdida_basada_en_minimo": "U", "perdida_basada_en_promedio": "V",
            "observaciones": "W",
        },
    },
    "ensayo_datos": {
        "data_start_row": 27,
        "mapping": {
            "id_punto": "A", "seg": "B", "cml": "C", "diametro_in": "D",
            "dja_mm": "E", "posicion_horario": "G", "longitud_mm": "I",
            "ancho_mm": "J", "espesor_minimo_medido_mm": "K",
            "interaccion_costura": "M", "tipo_anomalia": "O",
            "porcentaje_perdida": "Q", "observaciones": "S",
        },
    },
    "fotos": {"photo_row": 32, "desc_row": 33, "photo_cols": ["A", "G", "M", "S"]},
}

SCANC_RP_CONFIG = {
    "spreadsheet_id": "1Azx0v3Ur4oEockTDuC-FStnmK_HLLq6sY3FF9NePfg8",
    "template": "SCANC_RP.xlsx",
    "hoja_general": "1.0_general",
    "hoja_reporte_datos": "2.0_reporte",
    "hoja_ensayo_datos": "2.1_ensayo",
    "hoja_fotos": "2.0_reporte_photos",
    "celdas_generales": {
        "cliente": "D7", "fecha": "J7", "reporte_n": "Q7", "estacion": "V7",
        "contrato": "D8", "ot": "J8", "zona": "Q8", "sistema": "V8",
        "equipo": "D9", "fluido": "J9", "material": "Q9", "norma": "V9",
        "inicio_inspeccion": "D13", "acoplante": "J13", "estado_superficial": "R13",
        "marca_equipo": "V13", "fin_inspeccion": "D14", "rango_espesores": "J14",
        "temperatura_superficie": "R14", "modelo": "V14", "serie": "D15",
        "tipo_palpador": "J15", "frecuencia": "O15", "tamano": "R15", "fecha_calibracion": "V15",
        "nombre": "D36", "cargo": "D37", "certificado": "D38", "fecha_firma": "D40",
    },
    "celda_firma": "D39",
    "reporte_datos": {
        "data_start_row": 21,
        "mapping": {
            "id_punto": "A", "sistema_o_linea": "B", "cml": "D",
            "diametro_in": "E", "tipo_accesorio": "F", "dja_mm": "G",
            "posicion_horario_inicial": "J", "posicion_horario_final": "K",
            "longitud_barrido_circunferencial_mm": "L", "longitud_barrido_longitudinal_mm": "M",
            "area_barrido": "N", "numero_barridos": "O", "tipo_evaluacion": "P",
            "posicion_horario_soldadura_longitudinal": "Q", "espesor_nominal_mm": "R",
            "espesor_promedio_mm": "S", "espesor_minimo_mm": "T",
            "perdida_basada_en_minimo": "U", "perdida_basada_en_promedio": "V",
            "observaciones": "W",
        },
    },
    "ensayo_datos": {
        "data_start_row": 27,
        "mapping": {
            "id_punto": "A", "seg": "B", "cml": "C", "diametro_in": "D",
            "dja_mm": "E", "posicion_horario": "G", "longitud_mm": "I",
            "ancho_mm": "J", "espesor_minimo_medido_mm": "K",
            "interaccion_costura": "M", "tipo_anomalia": "O",
            "porcentaje_perdida": "Q", "observaciones": "S",
        },
    },
    "fotos": {"photo_row": 32, "desc_row": 33, "photo_cols": ["A", "G", "M", "S"]},
}


def _insertar_filas_y_ajustar_alturas(ws, pos: int, n: int):
    """Igual que en report_engine_570.py: openpyxl insert_rows() no desplaza
    alturas de fila ni merges — se hace manualmente aquí."""
    if n <= 0:
        return
    max_row_antes = ws.max_row
    alturas_originales = {
        r: ws.row_dimensions[r].height
        for r in range(pos, max_row_antes + 1)
        if r in ws.row_dimensions and ws.row_dimensions[r].height is not None
    }
    merges_a_mover = [r for r in list(ws.merged_cells.ranges) if r.min_row >= pos]
    for rng in merges_a_mover:
        ws.unmerge_cells(str(rng))

    ws.insert_rows(pos, n)

    for rng in merges_a_mover:
        rng.shift(0, n)
        ws.merge_cells(str(rng))

    for r in range(pos, max_row_antes + n + 1):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    for r_original, altura in alturas_originales.items():
        ws.row_dimensions[r_original + n].height = altura


def _replicar_merges_de_fila(ws, fila_origen: int, fila_destino: int):
    origen_merges = [
        rng for rng in list(ws.merged_cells.ranges)
        if rng.min_row == fila_origen and rng.max_row == fila_origen
    ]
    for rng in origen_merges:
        ref = (
            f"{get_column_letter(rng.min_col)}{fila_destino}:"
            f"{get_column_letter(rng.max_col)}{fila_destino}"
        )
        try:
            ws.merge_cells(ref)
        except ValueError:
            pass


def _copiar_estilo_fila(ws, fila_origen: int, fila_destino: int, max_col: int = 30):
    for c in range(1, max_col + 1):
        origen = ws.cell(row=fila_origen, column=c)
        destino = ws.cell(row=fila_destino, column=c)
        destino.font = copy(origen.font)
        destino.border = copy(origen.border)
        destino.fill = copy(origen.fill)
        destino.alignment = copy(origen.alignment)
        destino.number_format = origen.number_format


def _escribir_tabla(ws, seccion_cfg: dict, registros: list[dict], fila_inicio: int) -> int:
    """Escribe una tabla de datos de 2 filas de capacidad base, insertando
    filas extra si hace falta (igual patrón que report_engine_570.py).
    Devuelve cuántas filas extra se insertaron."""
    filas_extra = max(0, len(registros) - 2)
    if filas_extra > 0:
        fila_patron = fila_inicio + 1
        _insertar_filas_y_ajustar_alturas(ws, fila_patron + 1, filas_extra)
        altura_patron = ws.row_dimensions[fila_patron].height
        for i in range(filas_extra):
            fila_nueva = fila_patron + 1 + i
            _copiar_estilo_fila(ws, fila_patron, fila_nueva)
            _replicar_merges_de_fila(ws, fila_patron, fila_nueva)
            if altura_patron:
                ws.row_dimensions[fila_nueva].height = altura_patron

    for i, reg in enumerate(registros):
        fila_actual = fila_inicio + i
        for campo, col in seccion_cfg["mapping"].items():
            valor = reg.get(campo)
            if valor:
                ws[f"{col}{fila_actual}"] = valor_tipado(valor)

    return filas_extra


def generar_reporte_scanc(
    config: dict,
    fila_general: dict,
    reporte_datos: list[dict],
    ensayo_datos: list[dict],
    fotos: list[dict],
    progreso=None,
) -> bytes:
    """Genera el .xlsx real de SCAN C (líneas o RP, según `config`) y
    devuelve los bytes. `fotos` = lista de dicts {url, descripcion}."""
    def _reportar(pct: int, etapa: str):
        if progreso:
            progreso(pct, etapa)

    _reportar(3, "Preparando plantilla")
    wb = load_workbook(TEMPLATES_DIR / config["template"])
    ws = wb[HOJA_FORMATO]
    desactivar_fit_to_page(ws)

    _reportar(5, "Escribiendo datos generales")
    for campo, celda in config["celdas_generales"].items():
        valor = fila_general.get(campo)
        if valor:
            ws[celda] = valor_tipado(valor)

    filas_acumuladas = 0

    _reportar(15, "Tabla de reporte (escaneo)")
    fila_inicio = config["reporte_datos"]["data_start_row"] + filas_acumuladas
    filas_acumuladas += _escribir_tabla(ws, config["reporte_datos"], reporte_datos, fila_inicio)

    _reportar(35, "Tabla de información de ensayo")
    fila_inicio = config["ensayo_datos"]["data_start_row"] + filas_acumuladas
    filas_acumuladas += _escribir_tabla(ws, config["ensayo_datos"], ensayo_datos, fila_inicio)

    # ---- Fotos (mismo patrón de chunks de 4 que report_engine_570.py) ----
    fcfg = config["fotos"]
    base_photo_row = fcfg["photo_row"] + filas_acumuladas
    base_desc_row = fcfg["desc_row"] + filas_acumuladas

    chunks_necesarios = -(-max(len(fotos), 1) // 4)  # ceil(n/4), mínimo 1
    if chunks_necesarios > 1:
        filas_extra_fotos = (chunks_necesarios - 1) * 2
        _insertar_filas_y_ajustar_alturas(ws, base_photo_row + 1, filas_extra_fotos)
        for c in range(1, chunks_necesarios):
            f_foto = base_photo_row + c * 2
            f_desc = f_foto + 1
            _copiar_estilo_fila(ws, base_photo_row, f_foto)
            _copiar_estilo_fila(ws, base_desc_row, f_desc)
            _replicar_merges_de_fila(ws, base_photo_row, f_foto)
            _replicar_merges_de_fila(ws, base_desc_row, f_desc)
            ws.row_dimensions[f_foto].height = ws.row_dimensions[base_photo_row].height
            ws.row_dimensions[f_desc].height = ws.row_dimensions[base_desc_row].height
        filas_acumuladas += filas_extra_fotos

    total_fotos = len(fotos) or 1
    for i, foto in enumerate(fotos):
        chunk_idx = i // 4
        pos_in_chunk = i % 4
        f_foto = base_photo_row + chunk_idx * 2
        f_desc = f_foto + 1
        col = fcfg["photo_cols"][pos_in_chunk]

        desc = foto.get("descripcion") or ""
        if desc:
            ws[f"{col}{f_desc}"] = desc

        img_bytes = descargar_imagen(foto.get("url") or "")
        if img_bytes:
            insertar_imagen_centrada(ws, img_bytes, f"{col}{f_foto}")

        pct_fotos = 60 + round(((i + 1) / total_fotos) * 30)
        _reportar(min(pct_fotos, 90), f"Foto {i + 1} de {len(fotos)}")

    # ---- Firma (su fila real se corre hacia abajo por las filas insertadas
    # arriba, mismo caso que report_engine_570.py/510.py) ----
    _reportar(93, "Insertando firma")
    fila_firma_celda = config["celda_firma"]
    col_firma = fila_firma_celda[0]
    fila_firma = int(fila_firma_celda[1:]) + filas_acumuladas
    firma_bytes = descargar_imagen(fila_general.get("link_firma", ""))
    if firma_bytes:
        insertar_imagen_centrada(ws, firma_bytes, f"{col_firma}{fila_firma}")

    for campo in ("nombre", "cargo", "certificado", "fecha_firma"):
        celda = config["celdas_generales"].get(campo)
        if not celda:
            continue
        valor = fila_general.get(campo)
        if valor:
            col = "".join(ch for ch in celda if ch.isalpha())
            fila = int("".join(ch for ch in celda if ch.isdigit())) + filas_acumuladas
            ws[f"{col}{fila}"] = valor_tipado(valor)

    _reportar(97, "Guardando archivo")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
