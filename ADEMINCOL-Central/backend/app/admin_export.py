"""
Exportación de la base de datos a Excel para el rol ADMINISTRADOR
(2026-07-09). Cada tabla de Postgres se vuelve una HOJA del .xlsx; el admin
puede elegir qué tablas incluir y, dentro de cada una, filtrar/seleccionar
filas por su ID natural.

Registro CURADO (no se expone toda la BD a ciegas):
  - Solo las 7 tablas de negocio que el admin realmente administra (las
    tablas de sistema vacías —pmi_*, inspections, audit_log...— no aportan
    nada como descarga y solo confundirían).
  - `id_column`: la clave natural por la que se selecciona/filtra en el
    frontend (id_equipo, usuario, consecutivo...), NO el `id` SERIAL interno.
  - `exclude`: columnas que NUNCA salen en el Excel:
      * password_hash  -> seguridad, jamás debe salir de la BD.
      * firma_base64   -> string base64 de ~20 KB por fila; inflaría el
                          archivo y no sirve en una celda. Se reemplaza por
                          una columna "tiene_firma" (Sí/No) informativa.
"""
import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .db import fetch_all

# `grupo`: a qué categoría pertenece la tabla en el selector del frontend
# (2026-07-09, pedido explícito: "no como listas" — no todo junto en una
# sola lista plana). "General" son las tablas de soporte que ya tienen datos
# reales sincronizados desde Sheets (sync_service.py). El resto son grupos
# por técnica de reporte — hoy solo PMI tiene esquema en Postgres; cuando se
# migren SCAN C/570/510/Espesores, cada una suma su propio grupo aquí mismo,
# sin tocar el frontend (arma el selector dinámicamente a partir de esto).
EXPORTABLES: dict[str, dict] = {
    "users": {
        "label": "Usuarios",
        "grupo": "General",
        "id_column": "usuario",
        "exclude": ["password_hash", "firma_base64", "firma_mime"],
        "flag_firma": True,  # agrega columna "tiene_firma" derivada de firma_base64
    },
    "work_orders": {
        "label": "Órdenes de trabajo",
        "grupo": "General",
        "id_column": "id_ot",
        "exclude": [],
    },
    "servicios": {
        "label": "Servicios",
        "grupo": "General",
        "id_column": "id_servicio",
        "exclude": [],
    },
    "equipos_ensayo": {
        "label": "Equipos de ensayo",
        "grupo": "General",
        "id_column": "id_equipo",
        "exclude": [],
    },
    "personal_certificados": {
        "label": "Certificados del personal",
        "grupo": "General",
        "id_column": "id_certificado",
        "exclude": ["firma_base64"],  # se conserva firma_link (URL de origen)
        "flag_firma": True,
    },
    "certificados_usuarios": {
        "label": "Certificados de usuarios",
        "grupo": "General",
        "id_column": "id_certificado",
        "exclude": [],
    },
    "consecutivos_reportes": {
        "label": "Consecutivos de reportes",
        "grupo": "General",
        "id_column": "consecutivo",
        "exclude": [],
    },
    # PMI (Caracterización de Materiales) — esquema creado 2026-07-09
    # (ver backend/db/pmi_schema.sql), todavía sin datos migrados desde
    # Sheets (esa migración es un paso aparte, no incluido aquí). Se
    # agregan igual para que la pestaña quede completa desde ya.
    "pmi_general": {
        "label": "Datos generales",
        "grupo": "PMI",
        "id_column": "id_general",
        "exclude": [],
    },
    "pmi_quimica": {
        "label": "Química",
        "grupo": "PMI",
        "id_column": "id",
        "exclude": [],
    },
    "pmi_durezas": {
        "label": "Durezas",
        "grupo": "PMI",
        "id_column": "id",
        "exclude": [],
    },
}


# Caché en memoria de las columnas por tabla (2026-07-09): el esquema no
# cambia en caliente, así que consultar information_schema en CADA llamada
# es puro costo de red desperdiciado (~350-400ms por ida y vuelta a
# Supabase). Se resuelve una vez por tabla y queda fijo hasta que se
# reinicie el backend — si se altera una tabla, reiniciar para refrescar.
_cache_columnas: dict[str, list[str]] = {}


def _precargar_columnas():
    """Trae las columnas de las 7 tablas en UNA sola consulta
    (`table_name = ANY(...)`) y llena _cache_columnas de una vez — evita
    7 idas y vueltas separadas a Supabase en el primer uso."""
    faltantes = [k for k in EXPORTABLES if k not in _cache_columnas]
    if not faltantes:
        return
    filas = fetch_all(
        "SELECT table_name, column_name FROM information_schema.columns "
        "WHERE table_schema='public' AND table_name = ANY(%s) "
        "ORDER BY table_name, ordinal_position",
        (faltantes,),
    )
    por_tabla: dict[str, list[str]] = {k: [] for k in faltantes}
    for f in filas:
        por_tabla[f["table_name"]].append(f["column_name"])
    for key in faltantes:
        excl = set(EXPORTABLES[key].get("exclude", []))
        _cache_columnas[key] = [c for c in por_tabla[key] if c not in excl]


def _columnas_visibles(key: str) -> list[str]:
    """Columnas reales de la tabla, en orden, menos las excluidas."""
    if key not in _cache_columnas:
        _precargar_columnas()
    return _cache_columnas[key]


def listar_tablas() -> list[dict]:
    """Metadatos de cada tabla exportable, para que el frontend arme la UI
    sin hardcodear nada. Los conteos de las 7 tablas se piden en UNA sola
    ida y vuelta (UNION ALL) en vez de una consulta por tabla — con la red
    hacia Supabase, cada round-trip de más se nota (~350ms)."""
    keys = list(EXPORTABLES.keys())
    union_sql = " UNION ALL ".join(
        f"SELECT '{k}' AS tabla, count(*) AS c FROM \"{k}\"" for k in keys
    )
    conteos = {r["tabla"]: r["c"] for r in fetch_all(union_sql)}

    salida = []
    for key, cfg in EXPORTABLES.items():
        cols = _columnas_visibles(key)
        if cfg.get("flag_firma"):
            cols = cols + ["tiene_firma"]
        salida.append({
            "key": key,
            "label": cfg["label"],
            "grupo": cfg["grupo"],
            "idColumn": cfg["id_column"],
            "columnas": cols,
            "totalFilas": conteos.get(key, 0),
        })
    return salida


def _fila_normalizada(key: str, fila: dict, columnas: list[str]) -> dict:
    cfg = EXPORTABLES[key]
    out = {c: fila.get(c) for c in columnas if c != "tiene_firma"}
    if cfg.get("flag_firma"):
        # La firma en sí no se lee aquí (columnas ya la excluyó), pero para el
        # flag necesitamos saber si existe — se resuelve en la consulta con un
        # CASE, ver leer_filas().
        out["tiene_firma"] = fila.get("__tiene_firma")
    return out


def leer_filas(key: str) -> dict:
    """Todas las filas de una tabla (estas tablas son chicas, máx ~251
    filas — no se pagina). Devuelve columnas + filas listas para mostrar."""
    if key not in EXPORTABLES:
        raise KeyError(key)
    cfg = EXPORTABLES[key]
    columnas = _columnas_visibles(key)

    select_cols = ", ".join(f'"{c}"' for c in columnas)
    extra = ""
    if cfg.get("flag_firma"):
        col_firma = "firma_base64"
        extra = f", (CASE WHEN {col_firma} IS NOT NULL AND {col_firma} <> '' THEN 'Sí' ELSE 'No' END) AS __tiene_firma"

    filas = fetch_all(f'SELECT {select_cols}{extra} FROM "{key}" ORDER BY 1')

    cols_out = list(columnas)
    if cfg.get("flag_firma"):
        cols_out = cols_out + ["tiene_firma"]

    filas_norm = [_fila_normalizada(key, f, cols_out) for f in filas]
    return {
        "key": key,
        "label": cfg["label"],
        "idColumn": cfg["id_column"],
        "columnas": cols_out,
        "filas": filas_norm,
        "totalFilas": len(filas_norm),
    }


def _valor_celda(v):
    """openpyxl no acepta datetimes con timezone ('Excel does not support
    timezones'). Se limpia el tz; el resto se deja tal cual (openpyxl maneja
    date/número/bool nativamente)."""
    if isinstance(v, datetime) and v.tzinfo is not None:
        return v.replace(tzinfo=None)
    if isinstance(v, (dict, list)):
        # jsonb (ej. sync_runs.detalle, si algún día se expone) -> texto
        import json
        return json.dumps(v, ensure_ascii=False)
    return v


def generar_excel(seleccion: list[dict]) -> bytes:
    """`seleccion` = [{key, ids?}]. Por cada tabla se crea una hoja; si `ids`
    viene con valores, solo esas filas (por su id_column); si viene vacío o
    ausente, todas. Devuelve los bytes del .xlsx."""
    wb = Workbook()
    wb.remove(wb.active)  # quitar la hoja vacía por defecto

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C1121F")  # rojo ADEMINCOL

    hojas_creadas = 0
    for item in seleccion:
        key = item.get("key")
        if key not in EXPORTABLES:
            continue
        cfg = EXPORTABLES[key]
        datos = leer_filas(key)
        columnas = datos["columnas"]
        filas = datos["filas"]

        ids = set(str(x) for x in (item.get("ids") or []))
        if ids:
            id_col = cfg["id_column"]
            filas = [f for f in filas if str(f.get(id_col)) in ids]

        # Nombre de hoja: máx 31 chars, sin caracteres inválidos de Excel.
        nombre_hoja = cfg["label"][:31]
        for ch in "[]:*?/\\":
            nombre_hoja = nombre_hoja.replace(ch, " ")
        ws = wb.create_sheet(title=nombre_hoja)
        hojas_creadas += 1

        # Encabezado
        for c_idx, col in enumerate(columnas, start=1):
            celda = ws.cell(row=1, column=c_idx, value=col)
            celda.font = header_font
            celda.fill = header_fill

        # Filas
        for r_idx, fila in enumerate(filas, start=2):
            for c_idx, col in enumerate(columnas, start=1):
                ws.cell(row=r_idx, column=c_idx, value=_valor_celda(fila.get(col)))

        # Ancho de columnas aproximado (según el contenido más largo, con tope)
        for c_idx, col in enumerate(columnas, start=1):
            largo = len(str(col))
            for fila in filas:
                v = fila.get(col)
                if v is not None:
                    largo = max(largo, min(len(str(v)), 60))
            ws.column_dimensions[get_column_letter(c_idx)].width = min(largo + 2, 62)

        ws.freeze_panes = "A2"  # encabezado fijo al hacer scroll

    if hojas_creadas == 0:
        # Excel no permite un libro sin hojas.
        wb.create_sheet(title="Sin datos")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
