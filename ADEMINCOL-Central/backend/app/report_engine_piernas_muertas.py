"""
Generador REAL de reportes APP009 Piernas Muertas UT usando la plantilla
verificada (templates_xlsx/PIERNAS_MUERTAS.xlsx) y datos reales del Sheet
`ID_BD_GENERAL` (APP009_Piernas_Muertas_UT.js). Mismo patrón de cascada de
filas que report_engine_570.py, generalizado a 3 secciones (traducido 1:1 de
SECTIONS_CONFIG en el GAS original).

Diferencias deliberadas frente a 570/MT — el usuario pidió explícitamente
respetar la lógica original de este reporte, que es distinta:

1. **Sin firma.** El GAS original (`generarReporteUnico`) nunca escribe una
   firma/nombre/cargo en la plantilla — la sección "REALIZADO POR / REVISADO
   POR / APROBADO POR" (filas 52-57) queda tal cual el template la trae. No
   se replica aquí ninguna inserción de firma.
2. **Sin OT ni inspector en la hoja general.** `1_general` no tiene columnas
   de OT, inspector, cargo ni certificación — el reporte se genera por
   `id_pm` únicamente, jerarquizado por sistema (`0_sistema` → `1_general`).
3. **La sección "espesores" NO tiene bloque de fotos** (a diferencia de
   "inspeccion" y "radiografia") — así está en el GAS original, no es una
   omisión.
4. **Sin `link_reporte` en el Sheet.** El GAS original rastrea reportes ya
   generados listando archivos en una carpeta de Drive (`REPORTES_PIERNAS_
   MUERTAS/{id_sistema}_{nombreSistema}/`), no con una columna de link en la
   hoja. Aquí no replicamos ese rastreo por carpeta — el estado se reporta
   siempre como PENDIENTE (no hay forma de saber, solo con el Sheet, si ya
   se generó un reporte para un id_pm dado).
"""
import io
import logging
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .image_utils import desactivar_fit_to_page, descargar_imagen, insertar_imagen_centrada
from .report_utils import valor_tipado

logger = logging.getLogger("report_engine_piernas_muertas")

TEMPLATE_PATH = Path(__file__).resolve().parent / "templates_xlsx" / "PIERNAS_MUERTAS.xlsx"
HOJA_FORMATO = "formato"

# Traducido literal de MAPEO_DE_CELDAS en APP009_Piernas_Muertas_UT.js
CELDAS_GENERALES = {
    "cliente": "I8", "troncal": "V8", "estacion": "AI8", "sistema": "AV8",
    "nombre_pp": "V14", "fecha": "I10", "componente": "V10", "segmento": "AI10",
    "descripcion": "AI14", "inicio": "I12", "fin": "V12", "diametro": "AV12",
    "longitud": "AV10", "ref_linea asociada": "AI12", "orientacion": "I14",
}

# Traducido literal de SECTIONS_CONFIG en APP009_Piernas_Muertas_UT.js
# (2026-07-09). Verificado contra la plantilla real (filas 19/20, 39/40,
# 49/50 = 2 filas de capacidad cada una, igual que el patrón 570).
SECTIONS_CONFIG = {
    "inspeccion": {
        "sheet": "1_2_inspe_visual", "photo_sheet": "1_2_1_photos_vt", "data_start_row": 19,
        "mapping": {
            "recub_estado": "C", "recub_deterioro": "F", "valv_tipo": "I", "valv_rating_class": "K",
            "valv_humed": "N", "valv_corr": "Q", "valv_volante": "T", "brid_cumple_llen_tuer": "W",
            "brid_fugas": "AC", "rosca_fugas": "AF", "dano_mecanico": "AL", "prof_dano_mec": "AI",
            "prof_corr_ext": "AU", "corrosion_ext": "AP",
        },
        "photo_row": 23, "desc_row": 24, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "radiografia": {
        "sheet": "1_3radiografia", "photo_sheet": "1_3_1_photos_rt", "data_start_row": 39,
        "mapping": {
            "cml": "C", "componente": "F", "tiempo_exp_seg": "I", "iqi_obser": "L", "iqi_req": "O",
            "nps": "R", "thk_nom_mm": "U", "thk_min_mm": "X", "thk_prom_mm": "AA", "thk_corr_int_mm": "AD",
            "corr_interna": "AG", "thk_socavado_mm": "AJ", "thk_rosca_libre_mm": "AP",
            "indicaciones_soldaduras": "AM", "fluido": "AV", "valv_posicion": "AY",
            "objeto_interno": "BA", "sedimentos": "AS",
        },
        "photo_row": 43, "desc_row": 44, "photo_cols": ["C", "P", "AC", "AP"],
    },
    "espesores": {
        # Sin photo_sheet: el GAS original no tiene photosConfig para esta sección.
        "sheet": "1_1_med_espesores", "data_start_row": 49,
        "mapping": {
            "componente": "H", "nps": "G", "cml": "E",
            "med1": "K", "med2": "L", "med3": "M", "med4": "N", "med5": "O", "med6": "P", "med7": "Q",
            "med8": "R", "med9": "S", "med10": "T", "med11": "U", "med12": "V", "med13": "W", "med14": "X",
            "med15": "Y", "med16": "Z", "med17": "AA", "med18": "AB", "med19": "AC", "med20": "AD",
            "utc_min_1": "AG", "utc_prom_1": "AK", "utc_min_2": "AO", "utc_prom_2": "AS",
            "utc_min_3": "AW", "utc_prom_3": "AZ", "t_nominal": "AE",
        },
    },
}

SECTION_KEYS_ORDEN = list(SECTIONS_CONFIG.keys())


def _insertar_filas_y_ajustar_alturas(ws, pos: int, n: int):
    """Igual que en report_engine_570.py: openpyxl `insert_rows()` no
    desplaza alturas de fila ni merges — se hace manualmente aquí."""
    if n <= 0:
        return
    max_row_antes = ws.max_row
    alturas_originales = {
        r: ws.row_dimensions[r].height
        for r in range(pos, max_row_antes + 1)
        if r in ws.row_dimensions and ws.row_dimensions[r].height is not None
    }
    merges_a_mover = [r for r in list(ws.merged_cells.ranges) if r.min_row >= pos]
    for rng in merges_a_mover:
        ws.unmerge_cells(str(rng))

    ws.insert_rows(pos, n)

    for rng in merges_a_mover:
        rng.shift(0, n)
        ws.merge_cells(str(rng))

    for r in range(pos, max_row_antes + n + 1):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    for r_original, altura in alturas_originales.items():
        ws.row_dimensions[r_original + n].height = altura


def _replicar_merges_de_fila(ws, fila_origen: int, fila_destino: int):
    origen_merges = [
        rng for rng in list(ws.merged_cells.ranges)
        if rng.min_row == fila_origen and rng.max_row == fila_origen
    ]
    for rng in origen_merges:
        ref = (
            f"{get_column_letter(rng.min_col)}{fila_destino}:"
            f"{get_column_letter(rng.max_col)}{fila_destino}"
        )
        try:
            ws.merge_cells(ref)
        except ValueError:
            pass


def _copiar_estilo_fila(ws, fila_origen: int, fila_destino: int, max_col: int = 62):
    for c in range(1, max_col + 1):
        origen = ws.cell(row=fila_origen, column=c)
        destino = ws.cell(row=fila_destino, column=c)
        destino.font = copy(origen.font)
        destino.border = copy(origen.border)
        destino.fill = copy(origen.fill)
        destino.alignment = copy(origen.alignment)
        destino.number_format = origen.number_format


def generar_reporte_piernas_muertas(
    fila_general: dict,
    secciones_data: dict[str, list[dict]],
    secciones_fotos: dict[str, list[dict]],
    progreso=None,
) -> bytes:
    """Genera el .xlsx real de Piernas Muertas UT y devuelve los bytes.

    `secciones_data[key]` = lista de dicts (filas de la hoja de esa sección
    filtradas por id_pm). `secciones_fotos[key]` = lista de dicts
    {url, descripcion} para esa sección (vacío/ignorado para "espesores",
    que no tiene bloque de fotos en el original).
    """
    def _reportar(pct: int, etapa: str):
        if progreso:
            progreso(pct, etapa)

    _reportar(3, "Preparando plantilla")
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[HOJA_FORMATO]
    desactivar_fit_to_page(ws)

    _reportar(5, "Escribiendo datos generales")
    for campo, celda in CELDAS_GENERALES.items():
        valor = fila_general.get(campo)
        if valor:
            ws[celda] = valor_tipado(valor)

    total_fotos = sum(len(f) for f in secciones_fotos.values() if f) or 1
    fotos_procesadas = 0
    filas_acumuladas = 0

    for idx_sec, key in enumerate(SECTION_KEYS_ORDEN):
        config = SECTIONS_CONFIG[key]
        registros = secciones_data.get(key, [])
        tiene_fotos = "photo_sheet" in config
        fotos = secciones_fotos.get(key, []) if tiene_fotos else []

        pct_base = 10 + round((idx_sec / len(SECTION_KEYS_ORDEN)) * 80)
        _reportar(pct_base, f"Sección {idx_sec + 1}/{len(SECTION_KEYS_ORDEN)}: {config['sheet']}")

        fila_inicio = config["data_start_row"] + filas_acumuladas
        filas_extra_datos = max(0, len(registros) - 2)  # plantilla trae 2 filas de capacidad

        if filas_extra_datos > 0:
            fila_patron = fila_inicio + 1
            _insertar_filas_y_ajustar_alturas(ws, fila_patron + 1, filas_extra_datos)
            altura_patron = ws.row_dimensions[fila_patron].height
            for i in range(filas_extra_datos):
                fila_nueva = fila_patron + 1 + i
                _copiar_estilo_fila(ws, fila_patron, fila_nueva)
                _replicar_merges_de_fila(ws, fila_patron, fila_nueva)
                if altura_patron:
                    ws.row_dimensions[fila_nueva].height = altura_patron

        for i, reg in enumerate(registros):
            fila_actual = fila_inicio + i
            for campo, col in config["mapping"].items():
                valor = reg.get(campo)
                if valor:
                    ws[f"{col}{fila_actual}"] = valor_tipado(valor)

        filas_acumuladas += filas_extra_datos

        if not tiene_fotos:
            continue

        # ---- Fotos de esta sección ----
        base_photo_row = config["photo_row"] + filas_acumuladas
        base_desc_row = config["desc_row"] + filas_acumuladas

        chunks_necesarios = -(-max(len(fotos), 1) // 4)  # ceil(n/4), mínimo 1
        if chunks_necesarios > 1:
            filas_extra_fotos = (chunks_necesarios - 1) * 2
            _insertar_filas_y_ajustar_alturas(ws, base_photo_row + 1, filas_extra_fotos)
            for c in range(1, chunks_necesarios):
                f_foto = base_photo_row + c * 2
                f_desc = f_foto + 1
                _copiar_estilo_fila(ws, base_photo_row, f_foto)
                _copiar_estilo_fila(ws, base_desc_row, f_desc)
                _replicar_merges_de_fila(ws, base_photo_row, f_foto)
                _replicar_merges_de_fila(ws, base_desc_row, f_desc)
                ws.row_dimensions[f_foto].height = ws.row_dimensions[base_photo_row].height
                ws.row_dimensions[f_desc].height = ws.row_dimensions[base_desc_row].height
            filas_acumuladas += filas_extra_fotos

        for i, foto in enumerate(fotos):
            chunk_idx = i // 4
            pos_in_chunk = i % 4
            f_foto = base_photo_row + chunk_idx * 2
            f_desc = f_foto + 1
            col = config["photo_cols"][pos_in_chunk]

            desc = foto.get("descripcion") or ""
            if desc:
                ws[f"{col}{f_desc}"] = desc

            img_bytes = descargar_imagen(foto.get("url") or "")
            if img_bytes:
                insertar_imagen_centrada(ws, img_bytes, f"{col}{f_foto}")

            fotos_procesadas += 1
            pct_fotos = 10 + round((fotos_procesadas / total_fotos) * 80)
            _reportar(min(pct_fotos, 90), f"Foto {fotos_procesadas} de {total_fotos}")

    # Sin sección de firma: el GAS original no la escribe (ver docstring).

    _reportar(97, "Guardando archivo")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
