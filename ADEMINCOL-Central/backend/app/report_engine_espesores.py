"""
Generador REAL de reportes de Medición de Espesores (UT) usando la plantilla
verificada (templates_xlsx/ESPESORES.xlsx) y datos reales del Sheet
`DB_INSP_Medicion_Espesores`. Traducción de
ADEMINCOL-Scripts/.../Reporte_Medicion_Espesores.gs, verificada celda por
celda contra la plantilla real el 2026-07-09 — el script GAS pegado por el
usuario tenía DOS mapeos de celda incorrectos y uno vestigial, corregidos
aquí (ver comentarios junto a CELDAS_GENERALES).

Estructura: UNA sola tabla dinámica de lecturas (a diferencia de 570/510, que
tienen 15/11 secciones), con capacidad nativa de 2 filas (34-35, igual que
MT) y CINCO columnas con fórmulas vivas (MÁXIMO/MÍNIMO/PROMEDIO/%PÉRDIDA) que
hay que propagar manualmente a cada fila insertada — la única plantilla de
las 5 con este requisito, porque es la única cuya tabla depende de fórmulas
por fila en vez de solo datos.
"""
import io
import logging
import re
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .image_utils import (
    desactivar_fit_to_page,
    descargar_imagen,
    insertar_imagen_centrada,
    marcar_celda_sin_imagen,
)
from .report_utils import valor_tipado

logger = logging.getLogger("report_engine_espesores")

TEMPLATE_PATH = Path(__file__).resolve().parent / "templates_xlsx" / "ESPESORES.xlsx"
HOJA_FORMATO = "FORMATO"

# Verificado celda por celda contra la plantilla real el 2026-07-09 (dump de
# ws.merged_cells + texto de cada celda ancla). Dos correcciones sobre el
# script GAS pegado por el usuario:
#   - bloque_calibracion: el GAS decía 'AE23', que es la celda de la
#     ETIQUETA "Bloque de Calibración" (AE23:AG23) — el valor real es
#     AI23:AM23. Escribir en AE23 sobreescribía el texto de la etiqueta.
#   - procedimiento: el GAS decía 'P25', que cae DENTRO del merge de la
#     ETIQUETA "Procedimiento" (N25:P25) sin ser su celda ancla — openpyxl
#     lo rechaza (MergedCell no-ancla). El valor real es R25:Y25.
# NOTA: todas estas celdas están en filas 7-25, ANTES de la tabla de
# lecturas (fila 34+) — no se ven afectadas por la inserción de filas de
# lecturas/fotos, así que se escriben tal cual, sin ajuste de offset.
CELDAS_GENERALES = {
    "cliente": "D7", "contrato": "K7", "fecha_reporte": "U7", "ot": "AD7", "num_reporte": "AK7",
    "zona": "D9", "estacion": "K9", "sistema": "U9", "alcance": "AD9",
    "norma_referencia": "F11", "criterio_aceptacion": "AB11",
    "material": "E15", "temperatura_servicio": "R15", "tipo_recubrimiento": "AB15", "condicion_recubrimiento": "AJ15",
    "rating_sistema": "E17", "presion_diseno": "S17", "mop": "Z17", "codigo_diseno": "AG17",
    "marca_equipo": "G21", "modelo_equipo": "X21", "serie_equipo": "AF21", "fecha_calibracion": "AL21",
    "tipo_palpador": "E23", "frecuencia": "R23", "tamano_diametro": "AB23",
    "bloque_calibracion": "AI23",  # CORREGIDO — GAS decía AE23 (celda de la etiqueta)
    "material_bloque": "E25",
    "procedimiento": "R25",  # CORREGIDO — GAS decía P25 (MergedCell no-ancla)
    "tecnica": "AC25", "velocidad_calibracion": "AL25",
}
# 'link_foto_equipo' -> 'D21' del GAS original se descarta por completo: D21
# cae dentro del merge A21:E21 de la ETIQUETA "Marca Equipo Medidor de
# Espesores" (fila de solo 24px de alto, ni cabría una foto), Y la columna
# 'link_foto_equipo' NI SIQUIERA EXISTE en la hoja real '1_general'
# (verificado 2026-07-09) — ese campo del GAS nunca tuvo efecto en producción.

# Bloque de firmas, fila 39 ("REALIZADO POR" / "REVISADO POR" / "APROBADO
# POR"), 3 columnas de igual estructura (FIRMA/NOMBRE/CARGO/CERT. N°/FECHA
# en las filas 40-44). A DIFERENCIA de CELDAS_GENERALES, estas filas están
# DEBAJO de la tabla de lecturas/fotos — su posición real se corre hacia
# abajo por cada fila que se haya insertado dinámicamente arriba (mismo bug
# ya encontrado y corregido en report_engine_570.py/510.py: escribir en una
# fila fija producía "MergedCell... read-only" en cuanto la tabla insertaba
# filas). Por eso se escriben al final, con el offset acumulado sumado a la
# fila — ver `filas_extra` en generar_reporte_espesores().
#
# Columna 1 (C40:C44) = inspector, tal como el GAS ya lo llenaba. Columna 2
# (P40:P44, "REVISADO POR") es NUEVA aquí: mismo patrón ya establecido en
# PMI (P223-226) — se llena con el supervisor autenticado en la webapp que
# genera el reporte, resuelto contra la BD de usuarios (main.py), NO con una
# columna del Sheet de Espesores (que no tiene ese dato). Columna 3
# (AC40:AC44, "APROBADO POR") se deja en blanco, igual que el GAS actual: no
# existe todavía un tercer rol de aprobación en la plataforma.
CELDA_FIRMA = "C40"
CELDAS_FIRMA_TEXTO = {"nombre": "C41", "cargo": "C42", "certificado": "C43", "fecha": "C44"}
CELDA_FIRMA_SUPERVISOR = "P40"
CELDA_NOMBRE_SUPERVISOR = "P41"
CELDA_CARGO_SUPERVISOR = "P42"
CELDA_CERTIFICADO_SUPERVISOR = "P43"
CELDA_FECHA_SUPERVISOR = "P44"

# Columna 3 (AC40:AC44, "APROBADO POR") — pedido explícito del usuario
# 2026-07-14, mismo patrón que la columna 2 (REVISADO POR): usuario
# registrado en la plataforma o datos manuales, resuelto en main.py
# (_generar_bytes_espesores) contra 'aprobador_*' en fila_general. La fecha
# es siempre automática (hoy).
CELDA_FIRMA_APROBADOR = "AC40"
CELDA_NOMBRE_APROBADOR = "AC41"
CELDA_CARGO_APROBADOR = "AC42"
CELDA_CERTIFICADO_APROBADOR = "AC43"
CELDA_FECHA_APROBADOR = "AC44"

# Tabla de lecturas: capacidad nativa de 2 filas (34-35, igual que MT) —
# confirmado por la fórmula de conteo AG31 (`COUNT(J34:Y35)`, fija a esas 2
# filas igual que en el GAS: no se actualiza al insertar filas, limitación
# ya existente que no se corrige aquí para no desviarse del comportamiento
# validado). Nombres de campo en minúscula: coinciden con los headers reales
# de '2_lecturas_tomadas' (read_sheet_as_dicts los normaliza a minúsculas).
FILA_INICIO_LECTURAS = 34
COLUMNAS_LECTURAS = {
    "item": "A", "componente": "B", "cml": "F", "diametro": "H", "t_nominal": "I",
    "med1": "J", "med2": "K", "med3": "L", "med4": "M", "med5": "N", "med6": "O",
    "med7": "P", "med8": "Q", "med9": "R", "med10": "S", "med11": "T", "med12": "U",
    "med13": "V", "med14": "W", "med15": "X", "med16": "Y",
    "observaciones": "AJ",
}
# MÁXIMO, MÍNIMO, PROMEDIO, %PÉRDIDA Vs NOMINAL, %PÉRDIDA Vs PROMEDIO — traen
# fórmula en la plantilla, se propagan con _copiar_formulas_lecturas().
COLUMNAS_FORMULA_LECTURAS = ["Z", "AB", "AD", "AF", "AH"]

# Fotos: bloques de 3 columnas (A/N/AA), fila de foto (alto 332px) + fila de
# descripción (alto 19.5px) por bloque — confirmado contra la plantilla.
FILA_INICIO_FOTOS = 37
COLUMNAS_FOTOS = ["A", "N", "AA"]


def _col_fila(celda: str) -> tuple[str, int]:
    """Separa una referencia de celda en (columna, fila) — soporta columnas
    de dos letras como 'AC' (bug encontrado 2026-07-14 en report_engine_570.py
    al agregar un bloque de firma en columna de dos letras: `celda[0]`/
    `celda[1:]` asume una sola letra de columna)."""
    m = re.match(r"([A-Z]+)(\d+)", celda)
    return m.group(1), int(m.group(2))


def _corregir_referencias_grafico(ws):
    """El gráfico de barras nativo de la plantilla (comparación de espesores
    medidos vs. nominal) referencia una hoja 'FORMATOS_SCAN_C' que NO existe
    en este workbook — residuo de haber copiado la plantilla desde el
    formato SCAN C, que sí tiene una hoja con ese nombre. Como Excel no
    puede resolver esa hoja, el gráfico sale vacío/roto (ejes en blanco o
    "None"), aunque los datos SÍ están, en la hoja real 'FORMATO'
    (X27:AB27/X28:AB28, fórmulas que traen los valores de la tabla de
    lecturas) — reportado por el usuario 2026-07-16 con una captura del
    gráfico roto. Se corrige el nombre de hoja en cada referencia de serie."""
    for chart in ws._charts:
        for serie in chart.series:
            for ref_holder in (serie.val, serie.cat):
                if ref_holder is None:
                    continue
                for ref_attr in ("numRef", "strRef"):
                    ref = getattr(ref_holder, ref_attr, None)
                    if ref is not None and ref.f and "FORMATOS_SCAN_C" in ref.f:
                        ref.f = ref.f.replace("FORMATOS_SCAN_C", ws.title)


def _ajustar_formula_por_fila(formula: str, fila_origen: int, fila_destino: int) -> str:
    """Traducción literal de ajustarFormulaPorFila() en
    Reporte_Medicion_Espesores.gs: reemplaza toda referencia de celda a
    fila_origen dentro de la fórmula por fila_destino (ej. "=MAX(J34:Y34)"
    con origen=34, destino=35 -> "=MAX(J35:Y35)")."""
    patron = re.compile(rf"(\$?[A-Z]+\$?)({fila_origen})(?![0-9])")
    return patron.sub(lambda m: f"{m.group(1)}{fila_destino}", formula)


def _copiar_formulas_lecturas(ws, fila_base: int, total_filas: int):
    """Traducción de copiarFormulasLecturas(): propaga las fórmulas de
    MÁXIMO/MÍNIMO/PROMEDIO/%PÉRDIDA desde la fila patrón (fila_base) hacia
    cada fila adicional. Se ejecuta con 2+ lecturas, incluso usando solo las
    2 filas nativas de la plantilla (34-35): la fila 35 de la plantilla trae
    una fórmula de %pérdida LIGERAMENTE distinta a la de la 34 (sin el
    `MAX(0, ...)` de seguridad) — igual que el GAS original, aquí se
    sobrescribe con el patrón de la fila 34 para que ambas queden
    consistentes."""
    if total_filas <= 1:
        return
    for col in COLUMNAS_FORMULA_LECTURAS:
        formula_origen = ws[f"{col}{fila_base}"].value
        if not formula_origen or not str(formula_origen).startswith("="):
            continue
        for i in range(1, total_filas):
            fila_destino = fila_base + i
            ws[f"{col}{fila_destino}"] = _ajustar_formula_por_fila(str(formula_origen), fila_base, fila_destino)


def _insertar_filas_y_ajustar_alturas(ws, pos: int, n: int):
    """Mismo patrón que report_engine_mt.py/570/510: `ws.insert_rows()` no
    desplaza `row_dimensions` (alturas) ni `ws.merged_cells.ranges` — se
    hace manualmente aquí."""
    if n <= 0:
        return
    max_row_antes = ws.max_row
    alturas_originales = {
        r: ws.row_dimensions[r].height
        for r in range(pos, max_row_antes + 1)
        if r in ws.row_dimensions and ws.row_dimensions[r].height is not None
    }
    merges_a_mover = [r for r in list(ws.merged_cells.ranges) if r.min_row >= pos]
    for rng in merges_a_mover:
        ws.unmerge_cells(str(rng))

    ws.insert_rows(pos, n)

    for rng in merges_a_mover:
        rng.shift(0, n)
        ws.merge_cells(str(rng))

    for r in range(pos, max_row_antes + n + 1):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    for r_original, altura in alturas_originales.items():
        ws.row_dimensions[r_original + n].height = altura


def _replicar_merges_de_fila(ws, fila_origen: int, fila_destino: int):
    origen_merges = [
        rng for rng in list(ws.merged_cells.ranges)
        if rng.min_row == fila_origen and rng.max_row == fila_origen
    ]
    for rng in origen_merges:
        ref = (
            f"{get_column_letter(rng.min_col)}{fila_destino}:"
            f"{get_column_letter(rng.max_col)}{fila_destino}"
        )
        try:
            ws.merge_cells(ref)
        except ValueError:
            pass


def _copiar_estilo_fila(ws, fila_origen: int, fila_destino: int, max_col: int = 40):
    for c in range(1, max_col + 1):
        origen = ws.cell(row=fila_origen, column=c)
        destino = ws.cell(row=fila_destino, column=c)
        destino.font = copy(origen.font)
        destino.border = copy(origen.border)
        destino.fill = copy(origen.fill)
        destino.alignment = copy(origen.alignment)
        destino.number_format = origen.number_format


def generar_reporte_espesores(
    fila_general: dict,
    lecturas: list[dict],
    fotos: list[dict],
    progreso=None,
) -> bytes:
    """Genera el .xlsx real de Medición de Espesores y devuelve los bytes."""
    def _reportar(pct: int, etapa: str):
        if progreso:
            progreso(pct, etapa)

    _reportar(3, "Preparando plantilla")
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[HOJA_FORMATO]
    desactivar_fit_to_page(ws)
    _corregir_referencias_grafico(ws)

    _reportar(8, "Escribiendo datos generales")
    for campo, celda in CELDAS_GENERALES.items():
        valor = fila_general.get(campo)
        if valor:
            ws[celda] = valor_tipado(valor)

    # ---- Tabla de lecturas: insertar filas si hacen falta, escribir datos y propagar fórmulas ----
    _reportar(15, "Insertando filas de lecturas")
    n_lecturas = len(lecturas)
    filas_extra = max(0, n_lecturas - 2)  # plantilla ya trae 2 filas de capacidad (34-35)
    if filas_extra > 0:
        fila_patron = FILA_INICIO_LECTURAS + 1  # fila 35, ya con formato
        _insertar_filas_y_ajustar_alturas(ws, fila_patron + 1, filas_extra)
        altura_patron = ws.row_dimensions[fila_patron].height
        for i in range(filas_extra):
            fila_nueva = fila_patron + 1 + i
            _copiar_estilo_fila(ws, fila_patron, fila_nueva)
            _replicar_merges_de_fila(ws, fila_patron, fila_nueva)
            if altura_patron:
                ws.row_dimensions[fila_nueva].height = altura_patron

    _reportar(25, "Escribiendo lecturas")
    for idx, lectura in enumerate(lecturas):
        fila_actual = FILA_INICIO_LECTURAS + idx
        for campo, col in COLUMNAS_LECTURAS.items():
            valor = lectura.get(campo)
            if valor not in (None, ""):
                ws[f"{col}{fila_actual}"] = valor_tipado(valor)

    _copiar_formulas_lecturas(ws, FILA_INICIO_LECTURAS, n_lecturas)

    # ---- Fotos: bloques de 3 columnas, insertar pares de filas extra si hacen falta ----
    fila_base_foto = FILA_INICIO_FOTOS + filas_extra
    fila_base_desc = fila_base_foto + 1

    _reportar(35, "Insertando filas de fotos")
    chunks_necesarios = -(-max(len(fotos), 1) // 3)  # ceil(n/3), mínimo 1
    if chunks_necesarios > 1:
        filas_extra_fotos = (chunks_necesarios - 1) * 2
        # Insertar DESPUÉS de fila_base_desc (no en fila_base_foto + 1, que
        # ES fila_base_desc): insertar justo EN la fila patrón de descripción
        # la deja "contaminada" (blanco, altura None) por el mismo mecanismo
        # de _insertar_filas_y_ajustar_alturas que desplaza su contenido real
        # una fila más abajo — y como más adelante se usa esa fila como
        # FUENTE de estilo/alto para los bloques nuevos, el resultado era que
        # la fila de descripción (altura 19.5 en la plantilla) terminaba
        # perdiendo su alto y su estilo. Insertando después de ella, tanto
        # fila_base_foto como fila_base_desc quedan intactas como fuente
        # (mismo principio que la tabla de lecturas, que inserta después de
        # su fila patrón, nunca en ella). Bug encontrado y corregido el
        # 2026-07-09; el mismo patrón (`base_photo_row + 1`) existe en
        # report_engine_570.py/510.py — no corregido ahí en esta tarea.
        _insertar_filas_y_ajustar_alturas(ws, fila_base_desc + 1, filas_extra_fotos)
        for c in range(1, chunks_necesarios):
            f_foto = fila_base_foto + c * 2
            f_desc = f_foto + 1
            _copiar_estilo_fila(ws, fila_base_foto, f_foto)
            _copiar_estilo_fila(ws, fila_base_desc, f_desc)
            _replicar_merges_de_fila(ws, fila_base_foto, f_foto)
            _replicar_merges_de_fila(ws, fila_base_desc, f_desc)
            ws.row_dimensions[f_foto].height = ws.row_dimensions[fila_base_foto].height
            ws.row_dimensions[f_desc].height = ws.row_dimensions[fila_base_desc].height
        filas_extra += filas_extra_fotos

    total_fotos = len(fotos) or 1
    # Celdas de foto SIN imagen (menos fotos que espacios en el último bloque
    # de 3, o descarga fallida) quedan marcadas con una diagonal — pedido
    # del usuario 2026-07-16, mismo estilo "sin dato" del diálogo de bordes
    # de Sheets/Excel.
    total_slots_fotos = chunks_necesarios * 3
    for i in range(total_slots_fotos):
        chunk_idx = i // 3
        pos_in_chunk = i % 3
        f_foto = fila_base_foto + chunk_idx * 2
        f_desc = f_foto + 1
        col = COLUMNAS_FOTOS[pos_in_chunk]

        if i >= len(fotos):
            marcar_celda_sin_imagen(ws, f"{col}{f_foto}")
            continue

        foto = fotos[i]
        desc = foto.get("descripcion") or ""
        if desc:
            ws[f"{col}{f_desc}"] = desc

        pct = 40 + round((i / total_fotos) * 45)
        _reportar(min(pct, 85), f"Foto {i + 1} de {len(fotos)}")
        img_bytes = descargar_imagen(foto.get("url") or "")
        if img_bytes:
            insertar_imagen_centrada(ws, img_bytes, f"{col}{f_foto}")
        else:
            marcar_celda_sin_imagen(ws, f"{col}{f_foto}")

    # ---- Firmas: fila 39-44, DEBAJO de la tabla de lecturas/fotos — su
    # posición real quedó desplazada por `filas_extra` (lecturas + fotos
    # insertadas arriba). "REALIZADO POR" = inspector (CELDAS_FIRMA_TEXTO +
    # CELDA_FIRMA). "REVISADO POR" = supervisor que generó el reporte.
    _reportar(88, "Escribiendo datos del inspector")
    for campo, celda in CELDAS_FIRMA_TEXTO.items():
        valor = fila_general.get(campo)
        if valor:
            col = celda[0]
            fila = int(celda[1:]) + filas_extra
            ws[f"{col}{fila}"] = valor_tipado(valor)

    _reportar(90, "Insertando firma del inspector")
    col_firma, fila_firma = CELDA_FIRMA[0], int(CELDA_FIRMA[1:]) + filas_extra
    firma_bytes = descargar_imagen(fila_general.get("link_firma", ""))
    if firma_bytes:
        insertar_imagen_centrada(ws, firma_bytes, f"{col_firma}{fila_firma}", recortar_contenido=True)

    supervisor_nombre = fila_general.get("supervisor_nombre")
    if supervisor_nombre:
        _reportar(95, "Escribiendo datos del supervisor")
        col_nom, fila_nom = CELDA_NOMBRE_SUPERVISOR[0], int(CELDA_NOMBRE_SUPERVISOR[1:]) + filas_extra
        ws[f"{col_nom}{fila_nom}"] = supervisor_nombre
        if fila_general.get("supervisor_cargo"):
            col_c, fila_c = CELDA_CARGO_SUPERVISOR[0], int(CELDA_CARGO_SUPERVISOR[1:]) + filas_extra
            ws[f"{col_c}{fila_c}"] = fila_general["supervisor_cargo"]
        if fila_general.get("supervisor_certificado"):
            col_ce, fila_ce = CELDA_CERTIFICADO_SUPERVISOR[0], int(CELDA_CERTIFICADO_SUPERVISOR[1:]) + filas_extra
            ws[f"{col_ce}{fila_ce}"] = fila_general["supervisor_certificado"]
        col_f, fila_f = CELDA_FECHA_SUPERVISOR[0], int(CELDA_FECHA_SUPERVISOR[1:]) + filas_extra
        ws[f"{col_f}{fila_f}"] = datetime.now().strftime("%Y-%m-%d")
        firma_supervisor_bytes = descargar_imagen(fila_general.get("supervisor_firma_link", ""))
        if firma_supervisor_bytes:
            col_fs, fila_fs = CELDA_FIRMA_SUPERVISOR[0], int(CELDA_FIRMA_SUPERVISOR[1:]) + filas_extra
            insertar_imagen_centrada(
                ws, firma_supervisor_bytes, f"{col_fs}{fila_fs}", recortar_contenido=True
            )

    aprobador_nombre = fila_general.get("aprobador_nombre")
    if aprobador_nombre:
        _reportar(96, "Escribiendo datos del aprobador")
        col_nom, fila_nom = _col_fila(CELDA_NOMBRE_APROBADOR)
        ws[f"{col_nom}{fila_nom + filas_extra}"] = aprobador_nombre
        if fila_general.get("aprobador_cargo"):
            col_c, fila_c = _col_fila(CELDA_CARGO_APROBADOR)
            ws[f"{col_c}{fila_c + filas_extra}"] = fila_general["aprobador_cargo"]
        if fila_general.get("aprobador_certificado"):
            col_ce, fila_ce = _col_fila(CELDA_CERTIFICADO_APROBADOR)
            ws[f"{col_ce}{fila_ce + filas_extra}"] = fila_general["aprobador_certificado"]
        col_f, fila_f = _col_fila(CELDA_FECHA_APROBADOR)
        ws[f"{col_f}{fila_f + filas_extra}"] = datetime.now().strftime("%Y-%m-%d")
        firma_aprobador_bytes = descargar_imagen(fila_general.get("aprobador_firma_link", ""))
        if firma_aprobador_bytes:
            col_fs, fila_fs = _col_fila(CELDA_FIRMA_APROBADOR)
            insertar_imagen_centrada(
                ws, firma_aprobador_bytes, f"{col_fs}{fila_fs + filas_extra}", recortar_contenido=True
            )

    _reportar(97, "Guardando archivo")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
