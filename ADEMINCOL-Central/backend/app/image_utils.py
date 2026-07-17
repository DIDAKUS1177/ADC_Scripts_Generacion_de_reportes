"""
Utilidades de imágenes compartidas entre los motores de reporte (MT, PMI...).
Extraído de report_engine_mt.py el 2026-07-03 para reutilizar en PMI sin
duplicar código — cualquier fix aquí (como el bug de columnas >Z corregido
abajo) beneficia a todos los motores por igual.
"""
import base64
import io
import logging
import re
from concurrent.futures import ThreadPoolExecutor, as_completed

import httpx
from PIL import Image as PILImage, ImageChops
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.pagebreak import Break

logger = logging.getLogger("image_utils")


def desactivar_fit_to_page(ws):
    """REVERTIDO el 2026-07-08 — ya NO hace nada (se deja como no-op en vez
    de borrarla para no tener que tocar los 3 motores que la llaman).

    Historia: el 2026-07-06 diagnostiqué que `fitToPage=True` (que traen
    PMI/570/510, no MT) hacía que las imágenes flotantes se vieran
    desbordadas al exportar a PDF, y esta función lo desactivaba. Esa
    prueba se hizo con LibreOffice --headless (no había Excel real a mano).
    El 2026-07-08 el usuario reportó, probando en Excel de verdad, que con
    `fitToPage=False` el reporte PMI paginaba a **126 páginas** al imprimir
    (Ctrl+P) — la plantilla es MUY ancha (34 columnas) y alta (895 filas), y
    sin ese ajuste Excel no comprime nada. Osea: el bug de imágenes
    desbordadas era un artefacto del renderizador de LibreOffice, no de
    Excel real — la plantilla siempre estuvo pensada para imprimirse con
    `fitToPage=True` (así la dejó quien la diseñó) y ESO es lo correcto en
    Excel. Se revierte a no tocar el valor de la plantilla."""
    return


def drive_url_a_descarga(url: str) -> str | None:
    if not url:
        return None
    url = url.strip()
    if not url:
        return None
    if "drive.google.com" in url:
        m = re.search(r"id=([^&]+)", url) or re.search(r"/d/([^/]+)", url)
        if m:
            return f"https://drive.google.com/uc?export=download&id={m.group(1)}"
    return url if url.startswith("http") else None


def descargar_imagen(url: str, client: httpx.Client | None = None) -> bytes | None:
    """`client` opcional: si se pasa un `httpx.Client` ya abierto (ver
    `precargar_fotos`), reutiliza sus conexiones (keep-alive) en vez de
    abrir una TCP+TLS nueva por cada foto — con 2000 fotos al mismo host
    (AppSheet), ese handshake repetido era buena parte del tiempo perdido
    incluso ya paralelizando las descargas (reportado por el usuario
    2026-07-16: con 16 hilos en paralelo pero sin reuso de conexión, la
    velocidad solo mejoraba ~2x en vez de acercarse a 16x)."""
    if not url:
        return None
    url = url.strip()
    if url.startswith("data:image/"):
        try:
            _, b64_data = url.split(",", 1)
            return base64.b64decode(b64_data)
        except Exception:
            logger.warning("No se pudo decodificar imagen base64")
            return None

    final_url = drive_url_a_descarga(url)
    if not final_url:
        return None
    try:
        getter = client.get if client is not None else httpx.get
        resp = getter(final_url, timeout=15, follow_redirects=True)
        if resp.status_code == 200 and resp.content:
            return resp.content
    except Exception:
        logger.warning("No se pudo descargar imagen: %s", url)
    return None


def precargar_fotos(urls: list[str], progreso=None, max_workers: int = 24) -> dict[str, bytes | None]:
    """Descarga muchas fotos EN PARALELO (con un `httpx.Client` compartido
    para reusar conexiones, ver `descargar_imagen`) y devuelve
    {url: bytes|None} para consultar en vez de llamar `descargar_imagen`
    una por una durante el armado del .xlsx — pedido del usuario
    2026-07-16: un reporte de 570 con 2000 fotos tardaba demasiado porque
    cada descarga (red a AppSheet, ~1-2s) se esperaba de forma secuencial
    antes de pasar a la siguiente. openpyxl no es thread-safe (no se puede
    insertar en el workbook desde varios hilos), así que la inserción sigue
    siendo secuencial — pero la parte lenta, la RED, sí se paraleliza aquí.
    `progreso(pct, etapa)` se llama según van terminando las descargas
    (0-100 dentro de esta fase).

    URLs vacías/repetidas se filtran antes de lanzar hilos (no vale la pena
    un hilo por nada, y una foto repetida no debe descargarse dos veces)."""
    urls_unicas = list({u.strip() for u in urls if u and u.strip()})
    resultado: dict[str, bytes | None] = {}
    if not urls_unicas:
        return resultado

    completadas = 0
    total = len(urls_unicas)
    limites = httpx.Limits(max_connections=max_workers * 2, max_keepalive_connections=max_workers)
    with httpx.Client(limits=limites) as client, ThreadPoolExecutor(max_workers=max_workers) as executor:
        futuros = {executor.submit(descargar_imagen, u, client): u for u in urls_unicas}
        for futuro in as_completed(futuros):
            url = futuros[futuro]
            try:
                resultado[url] = futuro.result()
            except Exception:
                logger.warning("Error descargando foto en paralelo: %s", url)
                resultado[url] = None
            completadas += 1
            if progreso:
                progreso(round(completadas / total * 100), f"Descargando fotos ({completadas}/{total})")
    return resultado


def _recortar_a_contenido(image_bytes: bytes) -> bytes:
    """Recorta los márgenes en blanco/transparentes alrededor del trazo real
    de una firma antes de insertarla — sin esto, `insertar_imagen_centrada`
    centra el RECTÁNGULO del archivo de imagen, pero si la firma trae
    márgenes asimétricos (foto de un papel firmado, captura de pantalla con
    espacio de sobra a un lado) el trazo visible queda descentrado aunque el
    archivo en sí esté "centrado" (reportado por el usuario 2026-07-16, con
    capturas donde las firmas de Revisor/Aprobador se veían pegadas a una
    esquina)."""
    try:
        img = PILImage.open(io.BytesIO(image_bytes))
        img.load()
        if img.mode == "RGBA" or (img.mode == "P" and "transparency" in img.info):
            alpha = img.convert("RGBA").split()[-1]
            bbox = alpha.getbbox()
        else:
            rgb = img.convert("RGB")
            fondo = PILImage.new("RGB", rgb.size, (255, 255, 255))
            diff = ImageChops.difference(rgb, fondo)
            bbox = diff.getbbox()
        if not bbox:
            return image_bytes
        recortada = img.crop(bbox)
        buf = io.BytesIO()
        recortada.save(buf, format="PNG")
        return buf.getvalue()
    except Exception:
        logger.warning("No se pudo recortar la imagen al contenido, se usa tal cual")
        return image_bytes


def _comprimir_imagen(image_bytes: bytes, img_pil: "PILImage.Image", ancho_final: int, alto_final: int) -> bytes:
    """Reescala y reencoda a JPEG antes de incrustar — las fotos de celular
    vienen en resolución completa (3000x4000px, varios MB) pero se muestran
    a ~200x150px en la celda; sin comprimir, un reporte de 2000 fotos podía
    pesar varios GB y tardar minutos solo en guardarse (reportado por el
    usuario 2026-07-16). Se deja el DOBLE del tamaño mostrado (no 1:1) para
    que se vea nítida si alguien hace zoom o imprime. Solo para fotos de
    inspección — NO se usa en firmas (ver `insertar_imagen_centrada`), que
    ya son pequeñas tras `_recortar_a_contenido` y donde perder nitidez del
    trazo importa más que el ahorro de tamaño."""
    try:
        limite_w = max(ancho_final * 2, 60)
        limite_h = max(alto_final * 2, 60)
        img = img_pil
        if img.width > limite_w or img.height > limite_h:
            img = img.copy()
            img.thumbnail((limite_w, limite_h), PILImage.LANCZOS)
        if img.mode in ("RGBA", "LA", "P"):
            rgba = img.convert("RGBA")
            fondo = PILImage.new("RGB", rgba.size, (255, 255, 255))
            fondo.paste(rgba, mask=rgba.split()[-1])
            img = fondo
        elif img.mode != "RGB":
            img = img.convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=82, optimize=True)
        comprimida = buf.getvalue()
        return comprimida if len(comprimida) < len(image_bytes) else image_bytes
    except Exception:
        logger.warning("No se pudo comprimir la imagen, se usa tal cual")
        return image_bytes


def _ancho_columna(ws, col_letra: str) -> float:
    """Ancho de columna (unidades de caracter) SIN mutar la hoja.

    `ws.column_dimensions[letra]` (acceso por índice) es un defaultdict: si
    la columna no tenía ancho explícito, CREA un ColumnDimension nuevo — y
    ese objeto nuevo trae width=13.0 por defecto (no None). O sea que la
    sola LECTURA del ancho vía `[...]` graba un ancho de 13 en columnas que
    debían quedarse con el ancho por defecto de la plantilla. Bug real,
    confirmado 2026-07-09 con un reporte generado por el usuario: la
    plantilla PMI no trae ancho explícito en C:AG y el .xlsx generado salía
    con esas 31 columnas fijadas en 13.0 — esto, no la fórmula de escalado
    de imagen, era la causa de "las columnas siguen aumentando su tamaño".
    `.get()` no dispara el auto-creado (no pasa por `__missing__`)."""
    dim = ws.column_dimensions.get(col_letra)
    return dim.width if dim and dim.width else 8.43


def _alto_fila(ws, fila_num: int) -> float:
    """Alto de fila (puntos) sin mutar la hoja — ver `_ancho_columna`."""
    dim = ws.row_dimensions.get(fila_num)
    return dim.height if dim and dim.height else 15


def insertar_imagen_centrada(ws, image_bytes: bytes, celda_ancla: str, recortar_contenido: bool = False):
    """Inserta una imagen flotante centrada en el área de la celda (o su rango
    combinado), replicando el fix ya aplicado en los scripts GAS.

    `recortar_contenido=True` (solo para firmas, ver `_recortar_a_contenido`)
    recorta los márgenes en blanco/transparentes ANTES de centrar — para
    fotos de inspección normales debe quedar en False, porque una foto puede
    tener legítimamente fondo claro (cielo, pared) que NO hay que recortar.

    NOTA: el cálculo de ancho de columna usaba `chr(64 + c)` que solo
    funciona hasta la columna Z — se reemplazó por `get_column_letter()`
    (openpyxl) que soporta columnas AA, AB... correctamente. Importa para
    PMI, cuya plantilla llega hasta la columna AH."""
    col_letra, fila_num = coordinate_from_string(celda_ancla)
    col_idx = column_index_from_string(col_letra)

    destino = (col_idx, fila_num, col_idx, fila_num)
    for rng in ws.merged_cells.ranges:
        if rng.min_col <= col_idx <= rng.max_col and rng.min_row <= fila_num <= rng.max_row:
            destino = (rng.min_col, rng.min_row, rng.max_col, rng.max_row)
            break

    min_col, min_row, max_col, max_row = destino
    # Conversión ancho de columna (unidades de caracter, fuente Calibri 11,
    # MDW=7) -> píxeles. FORMULA CORREGIDA 2026-07-05: usaba solo `width * 7`,
    # que subestima el ancho real por ~5px POR COLUMNA (el ancho default de
    # Excel, 8.43, da 64px documentados — la fórmula vieja daba 59px). En un
    # merge de muchas columnas (ej. el bloque de fotos de 570/510, 22
    # columnas) el déficit acumulado superaba 100px, y como la imagen se
    # escalaba para caber en esa área SUBESTIMADA, quedaba más pequeña de lo
    # que debía y desplazada hacia la izquierda dentro de la celda real —
    # visualmente "no centrada" y con la columna viéndose más ancha que la
    # imagen.
    ancho_area = sum(_ancho_columna(ws, get_column_letter(c)) * 7 + 5 for c in range(min_col, max_col + 1))
    alto_area = sum(_alto_fila(ws, r) for r in range(min_row, max_row + 1)) * 96 / 72

    if recortar_contenido:
        image_bytes = _recortar_a_contenido(image_bytes)
    try:
        img_pil = PILImage.open(io.BytesIO(image_bytes))
        img_pil.load()
    except Exception:
        logger.warning("Imagen inválida, se omite")
        return

    margen = 4
    escala = min(
        max(ancho_area - margen, 20) / img_pil.width,
        max(alto_area - margen, 20) / img_pil.height,
    )
    ancho_final = round(img_pil.width * escala)
    alto_final = round(img_pil.height * escala)

    if not recortar_contenido:
        image_bytes = _comprimir_imagen(image_bytes, img_pil, ancho_final, alto_final)

    try:
        img = XLImage(io.BytesIO(image_bytes))
    except Exception:
        logger.warning("Imagen inválida tras comprimir, se omite")
        return
    img.width = ancho_final
    img.height = alto_final

    offset_x = max(0, round((ancho_area - ancho_final) / 2))
    offset_y = max(0, round((alto_area - alto_final) / 2))

    marker = AnchorMarker(
        col=min_col - 1, colOff=pixels_to_EMU(offset_x),
        row=min_row - 1, rowOff=pixels_to_EMU(offset_y),
    )
    size = XDRPositiveSize2D(cx=pixels_to_EMU(ancho_final), cy=pixels_to_EMU(alto_final))
    img.anchor = OneCellAnchor(_from=marker, ext=size)

    ws.add_image(img)


def marcar_celda_sin_imagen(ws, celda_ancla: str):
    """Marca con una diagonal (mismo estilo "sin dato" del diálogo de bordes
    de Sheets/Excel, línea de esquina inferior-izquierda a superior-derecha)
    la celda donde debía ir una foto de inspección pero no hay imagen
    disponible — pedido del usuario 2026-07-16. Se aplica solo a la celda
    ancla: si es parte de un rango combinado, Excel dibuja la diagonal a lo
    largo de TODO el rectángulo combinado (toma el borde del cell top-left),
    igual que ya hace con la imagen en `insertar_imagen_centrada`.

    Preserva los bordes existentes (arriba/abajo/izq/der) de la celda —
    solo agrega la diagonal, no reemplaza el resto del estilo."""
    actual = ws[celda_ancla].border
    ws[celda_ancla].border = Border(
        left=actual.left, right=actual.right, top=actual.top, bottom=actual.bottom,
        diagonal=Side(style="thin", color="FF000000"),
        diagonalUp=True, diagonalDown=False,
    )


def _alto_pagina_imprimible_pt(ws) -> float:
    """Alto imprimible de una página (puntos), asumiendo A4 (`paperSize` no
    viene seteado en las plantillas — la organización opera en Colombia,
    donde A4 es el estándar; si algún día se necesita Carta, este es el
    único lugar que tocar)."""
    ALTO_A4_MM = 297.0
    alto_pagina_in = ALTO_A4_MM / 25.4
    m = ws.page_margins
    alto_margenes_in = (m.top or 0) + (m.bottom or 0) + (m.header or 0) + (m.footer or 0)
    return (alto_pagina_in - alto_margenes_in) * 72


def evitar_fotos_partidas_entre_paginas(ws):
    """Inserta saltos de página manuales para que ninguna foto quede
    partida a la mitad entre dos páginas impresas — reportado por el
    usuario 2026-07-16 con una captura mostrando justo eso: una cuadrícula
    de fotos cortada entre las páginas 2 y 3 (los números de página de
    Excel en Vista previa de salto de página, visibles cruzando las
    fotos).

    Causa: cada fila de foto (alta, ~252pt) + su fila de descripción
    inmediata debajo forman una sola unidad visual, pero Excel calcula
    los saltos de página automáticos por altura acumulada SIN saber que
    esas dos filas van juntas — con reportes de cientos de fotos, tarde o
    temprano un salto cae justo en medio de alguna.

    Simula la paginación automática de Excel fila por fila (misma lógica:
    acumular alturas hasta que no quepa la siguiente), pero cuando la fila
    que no cabe es la fila de FOTO de un par foto+descripción, mueve el
    salto a ANTES del par completo en vez de dejarlo caer entre las dos.
    Es una heurística (el alto real de página en Excel puede variar unos
    puntos por redondeo de fuente/DPI) — no garantiza pixel-perfect, pero
    evita el caso reportado en la inmensa mayoría de los casos."""
    filas_con_foto = {img.anchor._from.row + 1 for img in ws._images}
    if not filas_con_foto:
        return

    alto_pagina = _alto_pagina_imprimible_pt(ws)

    def _alto_fila_o_default(fila: int) -> float:
        dim = ws.row_dimensions.get(fila)
        return dim.height if dim and dim.height else 15.0

    acumulado = 0.0
    fila = 1
    max_row = ws.max_row
    while fila <= max_row:
        if fila in filas_con_foto:
            alto_par = _alto_fila_o_default(fila) + _alto_fila_o_default(fila + 1)
            if acumulado > 0 and acumulado + alto_par > alto_pagina:
                ws.row_breaks.append(Break(id=fila - 1))
                acumulado = 0.0
            acumulado += alto_par
            fila += 2
            continue
        alto = _alto_fila_o_default(fila)
        if acumulado > 0 and acumulado + alto > alto_pagina:
            ws.row_breaks.append(Break(id=fila - 1))
            acumulado = 0.0
        acumulado += alto
        fila += 1
