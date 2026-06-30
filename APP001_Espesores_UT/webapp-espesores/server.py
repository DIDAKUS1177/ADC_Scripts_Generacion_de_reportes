"""
server.py — Servidor local para desarrollo de webapp-espesores
@author Diego Alejandro Hernandez Blanco

Lee directamente del Excel local (DB_INSP_Medicion_Espesores.xlsx)
y expone la misma API que Code.gs para que Index.html funcione
sin necesidad de desplegar en Google Apps Script.

Modo de uso:
    python server.py
    Luego abrir: http://localhost:8787/Index.html
"""

import os
import re
import shutil
import tempfile
import traceback
from io import BytesIO
from datetime import datetime

from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS
import openpyxl

# ──────────────────────────────────────────────────────────
#  CONFIG
# ──────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(
    os.path.dirname(BASE_DIR),
    'DB_INSP_Medicion_Espesores.xlsx'
)
PORT = 8787

app = Flask(__name__, static_folder=BASE_DIR)
CORS(app)

# ──────────────────────────────────────────────────────────
#  CONSTANTES — espejo de Code.gs
# ──────────────────────────────────────────────────────────
MAPEO_GENERAL = {
    'cliente':'D7',  'contrato':'K7',  'fecha_reporte':'U7', 'ot':'AD7',  'num_reporte':'AK7',
    'zona':'D9',     'estacion':'K9',  'sistema':'U8',       'alcance':'AD9',
    'norma_referencia':'F11',          'criterio_aceptacion':'AB11',
    'material':'E15','temperatura_servicio':'R15',
    'tipo_recubrimiento':'AB15',       'condicion_recubrimiento':'AJ15',
    'rating_sistema':'E17','presion_diseno':'S17','mop':'Z17','codigo_diseno':'AG17',
    'marca_equipo':'G21','modelo_equipo':'X21','serie_equipo':'AF21','fecha_calibracion':'AL21',
    'tipo_palpador':'E23','frecuencia':'R23','tamano_diametro':'AB23','bloque_calibracion':'AE23',
    'material_bloque':'E25','procedimiento':'P25','tecnica':'AC25','velocidad_calibracion':'AL25',
    'nombre':'C41',  'cargo':'C42',    'certificado':'C43',  'fecha':'C44'
}

FILA_INICIO_LECTURAS = 34
COLUMNAS_LECTURAS = {
    'item':'A',  'cml':'F',   'componente':'B', 'diametro':'H', 't_nominal':'I',
    'med1':'J',  'med2':'K',  'med3':'L',  'med4':'M',
    'med5':'N',  'med6':'O',  'med7':'P',  'med8':'Q',
    'med9':'R',  'med10':'S', 'med11':'T', 'med12':'U',
    'med13':'V', 'med14':'W', 'med15':'X', 'med16':'Y',
    'observaciones':'AJ'
}

# ──────────────────────────────────────────────────────────
#  HELPERS
# ──────────────────────────────────────────────────────────
def load_wb(data_only=True):
    return openpyxl.load_workbook(EXCEL_PATH, data_only=data_only)

def fmt_date(v):
    if isinstance(v, datetime):
        return v.strftime('%d/%m/%Y')
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s == 'None' else s

def clean(v):
    """Convierte valores de celda Excel a string limpio.
    Los números sin decimales (123.0) se convierten a '123', no '123.0'.
    """
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return '' if s == 'None' else s

def sheet_headers(ws):
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    return [clean(h).lower() for h in row]

def find_col(headers, *keys):
    for k in keys:
        for i, h in enumerate(headers):
            if k in h:
                return i
    return -1

# ──────────────────────────────────────────────────────────
#  RUTAS ESTÁTICAS
# ──────────────────────────────────────────────────────────
@app.route('/')
@app.route('/Index.html')
def serve_index():
    return send_from_directory(BASE_DIR, 'Index.html')

# ──────────────────────────────────────────────────────────
#  API — espejo de Code.gs
# ──────────────────────────────────────────────────────────
@app.route('/api/<fn>', methods=['POST'])
def api(fn):
    body = request.get_json(silent=True) or {}
    args = body.get('args', [])
    try:
        handlers = {
            'validarUsuario':       h_validar_usuario,
            'getReportesExistentes':h_get_reportes,
            'getDatosFormulario':   h_get_datos_formulario,
            'getLecturasPorId':     h_get_lecturas,
            'guardarLecturas':      h_guardar_lecturas,
            'eliminarLecturaCML':   h_eliminar_lectura,
            'generarReporte':       h_generar_reporte,
        }
        if fn not in handlers:
            return jsonify({'error': f'Función "{fn}" no implementada localmente'}), 404
        return jsonify(handlers[fn](args))
    except Exception as e:
        print(f'[ERROR] {fn}: {e}')
        return jsonify({'error': str(e)}), 500


# ── validarUsuario ──────────────────────────────────────
def h_validar_usuario(args):
    username = args[0].strip().lower() if len(args) > 0 else ''
    password = args[1].strip()         if len(args) > 1 else ''

    wb = load_wb()
    ws = wb['5_login']
    hdrs = sheet_headers(ws)
    iu   = find_col(hdrs, 'usuario', 'user')
    ip   = find_col(hdrs, 'contra', 'password', 'pass')

    if iu >= 0 and ip >= 0:
        for row in ws.iter_rows(min_row=2, values_only=True):
            u = clean(row[iu]).lower() if iu < len(row) else ''
            p = clean(row[ip])         if ip < len(row) else ''
            if u == username and p == password:
                return {'ok': True, 'username': username}

    return {'ok': False}


# ── getReportesExistentes ────────────────────────────────
def h_get_reportes(_args):
    wb = load_wb()

    # IDs con lecturas
    ws_lec  = wb['2_lecturas_tomadas']
    hdrs_l  = sheet_headers(ws_lec)
    i_id_l  = find_col(hdrs_l, 'id_general')
    ids_lec = set()
    if i_id_l >= 0:
        for row in ws_lec.iter_rows(min_row=2, values_only=True):
            v = clean(row[i_id_l]) if i_id_l < len(row) else ''
            if v:
                ids_lec.add(v)

    ws   = wb['1_general']
    hdrs = sheet_headers(ws)

    i_id = find_col(hdrs, 'id_general')
    i_cl = find_col(hdrs, 'client')
    i_ct = find_col(hdrs, 'contrat')
    i_fe = find_col(hdrs, 'fecha_rep', 'fecha rep')
    i_nr = find_col(hdrs, 'num_rep', 'numero_rep', 'nro_rep')
    i_es = find_col(hdrs, 'estacion', 'estación')
    i_si = find_col(hdrs, 'sistema')
    i_lk = find_col(hdrs, 'linkreporte', 'link_rep')

    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        id_val = clean(row[i_id]) if i_id >= 0 and i_id < len(row) else ''
        if not id_val:
            continue
        result.append({
            'id_general':   id_val,
            'cliente':      clean(row[i_cl]) if i_cl >= 0 and i_cl < len(row) else '',
            'contrato':     clean(row[i_ct]) if i_ct >= 0 and i_ct < len(row) else '',
            'estacion':     clean(row[i_es]) if i_es >= 0 and i_es < len(row) else '',
            'sistema':      clean(row[i_si]) if i_si >= 0 and i_si < len(row) else '',
            'fecha':        fmt_date(row[i_fe]) if i_fe >= 0 and i_fe < len(row) else '',
            'num_reporte':  clean(row[i_nr]) if i_nr >= 0 and i_nr < len(row) else '',
            'urlReporte':   clean(row[i_lk]) if i_lk >= 0 and i_lk < len(row) else '',
            'has_lecturas': id_val in ids_lec,
        })

    return result


# ── getDatosFormulario ───────────────────────────────────
def h_get_datos_formulario(_args):
    wb    = load_wb()
    ws    = wb['4_complementos']
    hdrs  = sheet_headers(ws)

    i_comp = find_col(hdrs, 'complementos', 'componente', 'component', 'comp')
    i_nps  = find_col(hdrs, 'nps_in', 'nps', 'diametr', 'diam')

    complementos, diametros = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        c = clean(row[i_comp]) if i_comp >= 0 and i_comp < len(row) else ''
        d = clean(row[i_nps])  if i_nps  >= 0 and i_nps  < len(row) else ''
        if c and c not in complementos:
            complementos.append(c)
        if d and d not in diametros:
            diametros.append(d)

    # B36
    ws_b36 = wb['B36']
    b36 = []
    for row in ws_b36.iter_rows(values_only=True):
        b36.append([clean(v) for v in row])

    return {
        'catalogos': {
            'TUBERIAS': {
                'complementos': complementos,
                'diametros':    diametros
            },
            'TANQUES': {
                'complementos': ['VIROLA / SHELL','FONDO','TECHO FIJO','TECHO FLOTANTE',
                                 'ANULAR (ANILLO DE FONDO)','BOQUILLA / NOZZLE',
                                 'PASO DE HOMBRE / MANWAY','TUBERÍA DE SUCCIÓN',
                                 'TUBERÍA DE DESCARGA','ESCALERA'],
                'diametros':    ['10 ft','20 ft','30 ft','40 ft','50 ft','3 m','6 m','9 m','12 m']
            },
            'RECIPIENTES_A_PRESION': {
                'complementos': ['CUERPO / SHELL','CABEZA ELÍPTICA 2:1','CABEZA HEMISFÉRICA',
                                 'CABEZA TORISFÉRICA','BOQUILLA / NOZZLE',
                                 'PASO DE HOMBRE / MANWAY','SKIRT / FALDA DE SOPORTE',
                                 'SILLA / SADDLE'],
                'diametros':    ['12"','18"','24"','30"','36"','42"','48"',
                                 '54"','60"','72"','84"','96"']
            }
        },
        'b36': b36
    }


# ── getLecturasPorId ─────────────────────────────────────
def h_get_lecturas(args):
    id_general = args[0] if args else ''

    wb    = load_wb()
    ws    = wb['2_lecturas_tomadas']
    raw_h = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    hdrs  = [clean(h) for h in raw_h]  # preservar case para el frontend
    hdrs_l = [h.lower() for h in hdrs]
    i_id   = find_col(hdrs_l, 'id_general')

    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if i_id >= 0 and clean(row[i_id]) == str(id_general).strip():
            obj = {}
            for i, h in enumerate(hdrs):
                obj[h] = clean(row[i]) if i < len(row) else ''
            result.append(obj)

    return result


# ── guardarLecturas ──────────────────────────────────────
def h_guardar_lecturas(args):
    id_general = args[0] if len(args) > 0 else ''
    filas      = args[1] if len(args) > 1 else []

    if not id_general:
        return {'success': False, 'error': 'id_general requerido'}

    wb   = openpyxl.load_workbook(EXCEL_PATH)
    ws   = wb['2_lecturas_tomadas']
    raw_h = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    hdrs  = [clean(h) for h in raw_h]
    hdrs_l = [h.lower() for h in hdrs]
    i_id   = find_col(hdrs_l, 'id_general')

    # Borrar filas existentes para este id (de abajo hacia arriba)
    rows_to_del = [
        r for r in range(2, ws.max_row + 1)
        if i_id >= 0 and clean(ws.cell(r, i_id + 1).value) == str(id_general).strip()
    ]
    for r in reversed(rows_to_del):
        ws.delete_rows(r)

    # Insertar nuevas filas
    for fila in filas:
        new_row = []
        for h in hdrs_l:
            if h == 'id_general':
                new_row.append(id_general)
            elif h == 'item':
                new_row.append(fila.get('item', ''))
            elif h == 'cml':
                new_row.append(fila.get('CML', fila.get('cml', '')))
            elif h == 'componente':
                new_row.append(fila.get('componente', ''))
            elif h == 'diametro':
                new_row.append(fila.get('diametro', ''))
            elif h == 't_nominal':
                new_row.append(fila.get('t_nominal', ''))
            elif h == 'observaciones':
                new_row.append(fila.get('observaciones', ''))
            else:
                m = re.match(r'^m(?:ed)?0?(\d{1,2})$', h)
                if m:
                    new_row.append(fila.get('med' + str(int(m.group(1))), ''))
                else:
                    new_row.append('')
        ws.append(new_row)

    wb.save(EXCEL_PATH)
    return {'success': True, 'filas': len(filas), 'id': id_general}


# ── eliminarLecturaCML ───────────────────────────────────
def h_eliminar_lectura(args):
    id_general = args[0] if len(args) > 0 else ''
    cml_nombre = args[1] if len(args) > 1 else ''

    wb    = openpyxl.load_workbook(EXCEL_PATH)
    ws    = wb['2_lecturas_tomadas']
    hdrs  = sheet_headers(ws)
    i_id  = find_col(hdrs, 'id_general')
    i_cml = find_col(hdrs, 'cml')

    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        id_v  = clean(ws.cell(r, i_id  + 1).value) if i_id  >= 0 else ''
        cml_v = clean(ws.cell(r, i_cml + 1).value) if i_cml >= 0 else ''
        if id_v == str(id_general).strip() and cml_v == str(cml_nombre).strip():
            rows_to_del.append(r)

    for r in reversed(rows_to_del):
        ws.delete_rows(r)

    wb.save(EXCEL_PATH)
    return {'success': True, 'eliminadas': len(rows_to_del)}


# ── generarReporte (vía API — redirige al endpoint de descarga) ──
def h_generar_reporte(args):
    # En local, el frontend usa /generar/<id> directo; este endpoint queda como fallback
    id_general = args[0] if args else ''
    return {'success': False, 'error': 'Usa el endpoint /generar/' + id_general}


# ──────────────────────────────────────────────────────────
#  GENERADOR DE REPORTES LOCAL  (descarga XLSX directo)
# ──────────────────────────────────────────────────────────
@app.route('/generar/<path:id_general>', methods=['POST'])
def generar_reporte_local(id_general):
    id_general = id_general.strip()
    try:
        # 1 ── Leer datos generales
        wb_d      = load_wb()   # data_only=True — valores sin fórmulas
        ws_gen    = wb_d['1_general']
        gen_raw_h = next(ws_gen.iter_rows(min_row=1, max_row=1, values_only=True), [])
        gen_hdrs  = [clean(h).lower() for h in gen_raw_h]
        i_id      = find_col(gen_hdrs, 'id_general')

        datos_gral = {}
        for row in ws_gen.iter_rows(min_row=2, values_only=True):
            if i_id >= 0 and i_id < len(row) and clean(row[i_id]) == id_general:
                for i, h in enumerate(gen_hdrs):
                    datos_gral[h] = clean(row[i]) if i < len(row) else ''
                break

        if not datos_gral:
            return jsonify({'error': f'ID "{id_general}" no encontrado en 1_general'}), 404

        # 2 ── Leer lecturas
        ws_lec   = wb_d['2_lecturas_tomadas']
        lec_raw_h = next(ws_lec.iter_rows(min_row=1, max_row=1, values_only=True), [])
        lec_hdrs  = [clean(h).lower() for h in lec_raw_h]
        i_id_l    = find_col(lec_hdrs, 'id_general')

        lecturas = []
        for row in ws_lec.iter_rows(min_row=2, values_only=True):
            if i_id_l >= 0 and i_id_l < len(row) and clean(row[i_id_l]) == id_general:
                obj = {lec_hdrs[i]: (clean(row[i]) if i < len(row) else '') for i in range(len(lec_hdrs))}
                lecturas.append(obj)

        # 3 ── Copiar plantilla a archivo temporal
        tmp_path = os.path.join(tempfile.gettempdir(), f'rpt_{id_general[:30].replace("/","_")}.xlsx')
        shutil.copy2(EXCEL_PATH, tmp_path)

        # 4 ── Abrir copia y eliminar hojas que no son la plantilla
        wb_out = openpyxl.load_workbook(tmp_path)
        for s in [n for n in wb_out.sheetnames if n != 'FORMATOS_SCAN_C']:
            del wb_out[s]
        ws = wb_out['FORMATOS_SCAN_C']

        # 5 ── Escribir datos generales
        for campo, celda in MAPEO_GENERAL.items():
            val = datos_gral.get(campo, '')
            if val:
                try:
                    ws[celda] = val
                except Exception:
                    pass

        # 6 ── Insertar filas extra para lecturas adicionales
        n = len(lecturas)
        if n > 1:
            ws.insert_rows(FILA_INICIO_LECTURAS + 1, amount=n - 1)

        # 7 ── Escribir lecturas fila a fila
        for idx, lec in enumerate(lecturas):
            fila = FILA_INICIO_LECTURAS + idx
            for campo, col in COLUMNAS_LECTURAS.items():
                val = lec.get(campo, '')
                if val:
                    try:
                        ws[f'{col}{fila}'] = val
                    except Exception:
                        pass

        # 8 ── Exportar a BytesIO y devolver como descarga
        out = BytesIO()
        wb_out.save(out)
        out.seek(0)

        try:
            os.remove(tmp_path)
        except Exception:
            pass

        safe = id_general[:40].replace('/', '_').replace('\\', '_')
        return send_file(
            out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Reporte_Espesores_{safe}.xlsx'
        )

    except Exception as e:
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# ──────────────────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────────────────
if __name__ == '__main__':
    print(f'\n  ADEMINCOL · Servidor local de desarrollo')
    print(f'  Excel: {EXCEL_PATH}')
    print(f'  URL:   http://localhost:{PORT}/Index.html\n')
    app.run(port=PORT, debug=False)
